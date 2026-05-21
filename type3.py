import streamlit as st
import pandas as pd
import numpy as np
import re
import base64
import os
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


def normalize_arabic(text):
    if not isinstance(text, str):
        return text
    text = text.strip()
    text = re.sub(r'[أإآا]', 'ا', text)
    text = re.sub(r'[ىي]', 'ي', text)
    text = re.sub(r'[ةه]', 'ه', text)
    text = re.sub(r'[\u064B-\u065F]', '', text)
    return text


def has_star_rating(df):
    col = 'التقييم الداخلي/الأساسي'
    if col not in df.columns:
        return False
    return df[col].astype(str).str.strip().eq('*').any()


def process_airport_data(df):
    num_cols = [
        'Domestic Arrival', 'Domestic Departure',
        'International Arrival Total', 'International Departure Total'
    ]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    grouped = df.groupby('Airport')[num_cols].sum().reset_index()
    grouped.columns = [
        'Airport',
        'Total of Domestic Arrival', 'Total of Domestic Departure',
        'Total of International Arrival', 'Total of International Departure'
    ]
    grouped['Total of Domestic']      = grouped['Total of Domestic Arrival']      + grouped['Total of Domestic Departure']
    grouped['Total of International'] = grouped['Total of International Arrival'] + grouped['Total of International Departure']
    grouped['total pas']              = grouped['Total of Domestic']              + grouped['Total of International']
    grouped['use it in for (DOM)']    = grouped.apply(
        lambda r: r['Total of Domestic']      / r['total pas'] if r['total pas'] > 0 else 0, axis=1)
    grouped['use it in for (DOM)2']   = grouped.apply(
        lambda r: r['Total of International'] / r['total pas'] if r['total pas'] > 0 else 0, axis=1)
    return grouped


def set_background():
    for ext in ['png', 'jpg', 'jpeg', 'webp']:
        path = f"Background.{ext}"
        if os.path.exists(path):
            with open(path, "rb") as f:
                data = base64.b64encode(f.read()).decode()
            st.markdown(f"""
                <style>
                .stApp {{
                    background-image: url("data:image/{ext};base64,{data}");
                    background-size: cover;
                    background-position: center;
                    background-attachment: fixed;
                }}
                .block-container {{
                    background-color: rgba(255, 255, 255, 0.82);
                    border-radius: 12px;
                    padding: 1.5rem;
                }}
                .stButton > button[kind="primary"] {{
                    background-color: #003366 !important;
                    border-color: #003366 !important;
                    color: white !important;
                }}
                .stButton > button[kind="primary"]:hover {{
                    background-color: #00509e !important;
                    border-color: #00509e !important;
                }}
                </style>
            """, unsafe_allow_html=True)
            return


# ─── ثوابت المطارات ───
AIRPORT_FULL_NAMES = {
    "ABH": "Abha International Airport",
    "GIZ": "King Abdullah bin Abdulaziz International Airport - Jizan",
    "TUU": "Prince Sultan bin Abdulaziz International Airport - Tabuk",
}
AIRPORT_LABELS = {
    "ABH": "🛬 ABH - مطار أبها",
    "GIZ": "🛬 GIZ - مطار جيزان",
    "TUU": "🛬 TUU - مطار تبوك",
}


# =============================================
# دوال Excel
# =============================================
def _excel_styles():
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    alt_fill     = PatternFill("solid", fgColor="D6E4F0")
    white_fill   = PatternFill("solid", fgColor="FFFFFF")
    green_fill   = PatternFill("solid", fgColor="E2EFDA")
    bold_white   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font  = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin',  color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
    )
    return dict(
        header_fill=header_fill, alt_fill=alt_fill, white_fill=white_fill,
        green_fill=green_fill, bold_white=bold_white, normal_font=normal_font,
        center_align=center_align, thin_border=thin_border
    )


def _sh(ws, row, cols, s):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = s['header_fill']; cell.font = s['bold_white']
        cell.alignment = s['center_align']; cell.border = s['thin_border']


def _sd(ws, row, cols, s, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = s['alt_fill'] if alt else s['white_fill']
        cell.font = s['normal_font']
        cell.alignment = s['center_align']; cell.border = s['thin_border']


def _write_summary_sheet(wb, results_table, s, type_label="Type 3"):
    ws1 = wb.active
    ws1.title = "النتائج التراكمية"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells("A1:C1")
    ws1["A1"] = f"الجدول التراكمي النهائي - {type_label}"
    ws1["A1"].font      = Font(name="Arial", bold=True, color="1F4E79", size=14)
    ws1["A1"].alignment = s['center_align']
    ws1["A1"].fill      = PatternFill("solid", fgColor="D6E4F0")
    ws1.row_dimensions[1].height = 30
    for i, h in enumerate(["المطار", "النتيجة الخام", "النتيجة النهائية (÷0.85)"], 1):
        ws1.cell(row=2, column=i, value=h)
    _sh(ws1, 2, 3, s)
    for r_idx, row in enumerate(results_table, start=3):
        ws1.cell(row=r_idx, column=1, value=row["المطار"])
        ws1.cell(row=r_idx, column=2, value=row["النتيجة الخام"])
        ws1.cell(row=r_idx, column=3, value=row["النتيجة النهائية (÷0.85)"])
        _sd(ws1, r_idx, 3, s, alt=(r_idx % 2 == 0))
        ws1.cell(row=r_idx, column=2).number_format = "0.0000%"
        ws1.cell(row=r_idx, column=3).number_format = "0.0000%"
        ws1.cell(row=r_idx, column=3).fill = s['green_fill']
        ws1.cell(row=r_idx, column=3).font = Font(name="Arial", bold=True, color="375623", size=10)
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 28


def _write_detail_sheets(wb, detail_data, s):
    for sheet_name, df_detail in detail_data.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.rightToLeft = True
        for c_idx, col_name in enumerate(df_detail.columns, 1):
            ws.cell(row=1, column=c_idx, value=col_name)
        _sh(ws, 1, len(df_detail.columns), s)
        for r_idx, row in enumerate(df_detail.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            _sd(ws, r_idx, len(df_detail.columns), s, alt=(r_idx % 2 == 0))
        for c_idx, col_name in enumerate(df_detail.columns, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = max(18, len(str(col_name)) + 4)


def build_excel_report(results_table, detail_data=None):
    s  = _excel_styles()
    wb = Workbook()
    _write_summary_sheet(wb, results_table, s)
    if detail_data:
        _write_detail_sheets(wb, detail_data, s)
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out


def fix_arabic(text):
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        return get_display(arabic_reshaper.reshape(str(text)))
    except Exception:
        return str(text)


def make_chart_main(df_main, airport_code):
    categories = [fix_arabic(c) for c in df_main['Category'].tolist()]
    values     = [float(v) * 100 for v in df_main['النتيجة النهائية'].tolist()]
    fig, ax = plt.subplots(figsize=(13, max(5, len(categories) * 0.55 + 2)))
    fig.patch.set_facecolor('#F7FBFF')
    ax.set_facecolor('#F7FBFF')
    colors = ['#1F4E79' if v >= 70 else '#2E75B6' if v >= 50 else '#9DC3E6' for v in values]
    bars = ax.barh(categories, values, color=colors, edgecolor='white', height=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.2, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va='center', ha='left', fontsize=9, color='#1F4E79', fontweight='bold')
    ax.set_xlim(0, max(values) * 1.3 if values else 100)
    ax.set_xlabel(fix_arabic('النتيجة النهائية'), fontsize=10, color='#1F4E79')
    ax.set_title(fix_arabic(f'نتائج التصنيفات الرئيسية - {airport_code}'),
                 fontsize=13, fontweight='bold', color='#1F4E79', pad=15)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#BFBFBF'); ax.spines['bottom'].set_color('#BFBFBF')
    ax.grid(axis='x', linestyle='--', alpha=0.4, color='#BFBFBF')
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='#1F4E79', label=fix_arabic('ممتاز (≥ 70%)')),
        Patch(facecolor='#2E75B6', label=fix_arabic('جيد (≥ 50%)')),
        Patch(facecolor='#9DC3E6', label=fix_arabic('يحتاج تحسين (< 50%)')),
    ]
    ax.legend(handles=legend_elements, loc='lower right', fontsize=8, framealpha=0.7, edgecolor='#BFBFBF')
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); buf.seek(0)
    return buf


def make_chart_sub_by_category(df_sub, airport_code, category_name):
    df_cat = df_sub[df_sub['Category'] == category_name].copy()
    if df_cat.empty:
        return None
    labels = [fix_arabic(r['Subcategory']) for _, r in df_cat.iterrows()]
    values = [float(v) * 100 for v in df_cat['الوزن × الإنجاز'].tolist()]
    fig, ax = plt.subplots(figsize=(12, max(3.5, len(labels) * 0.5 + 1.5)))
    fig.patch.set_facecolor('#F7FBFF'); ax.set_facecolor('#F7FBFF')
    palette = plt.cm.Blues(np.linspace(0.45, 0.88, max(len(labels), 1)))
    bars = ax.barh(labels, values, color=palette, edgecolor='white', height=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va='center', ha='left', fontsize=9, color='#1F4E79')
    ax.set_xlim(0, max(values) * 1.3 if values else 100)
    ax.set_xlabel(fix_arabic('الوزن × الإنجاز'), fontsize=9, color='#1F4E79')
    ax.set_title(fix_arabic(f'{category_name} | {airport_code}'),
                 fontsize=11, fontweight='bold', color='#1F4E79', pad=10)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#BFBFBF'); ax.spines['bottom'].set_color('#BFBFBF')
    ax.grid(axis='x', linestyle='--', alpha=0.4, color='#BFBFBF')
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); buf.seek(0)
    return buf


def build_excel_with_charts(results_table, detail_data):
    s  = _excel_styles()
    wb = Workbook()
    _write_summary_sheet(wb, results_table, s)
    if detail_data:
        _write_detail_sheets(wb, detail_data, s)

    airports_in_data = {k.replace("MainCat_", "") for k in detail_data if k.startswith("MainCat_")}
    if airports_in_data:
        ws_c = wb.create_sheet(title="الرسومات البيانية")
        ws_c.sheet_view.rightToLeft = True
        ws_c["A1"] = "الرسومات البيانية"
        ws_c["A1"].font      = Font(name="Arial", bold=True, color="1F4E79", size=14)
        ws_c["A1"].fill      = PatternFill("solid", fgColor="D6E4F0")
        ws_c["A1"].alignment = s['center_align']
        ws_c.row_dimensions[1].height = 30
        cur_row = 3

        for code in sorted(airports_in_data):
            ws_c.cell(row=cur_row, column=1, value=f"مطار {code}")
            ws_c.cell(row=cur_row, column=1).font      = Font(name="Arial", bold=True, color="FFFFFF", size=13)
            ws_c.cell(row=cur_row, column=1).fill      = PatternFill("solid", fgColor="1F4E79")
            ws_c.cell(row=cur_row, column=1).alignment = s['center_align']
            ws_c.merge_cells(f"A{cur_row}:P{cur_row}")
            ws_c.row_dimensions[cur_row].height = 28
            cur_row += 1

            if f"MainCat_{code}" in detail_data:
                df_m  = detail_data[f"MainCat_{code}"]
                buf_m = make_chart_main(df_m, code)
                img_m = XLImage(buf_m)
                img_m.width  = 750
                img_m.height = max(320, len(df_m) * 28 + 80)
                ws_c.add_image(img_m, f"A{cur_row}")
                cur_row += max(22, int(img_m.height / 15) + 2)

            if f"SubCat_{code}" in detail_data:
                df_s = detail_data[f"SubCat_{code}"]
                ws_c.cell(row=cur_row, column=1, value=f"التصنيفات الفرعية - {code}")
                ws_c.cell(row=cur_row, column=1).font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                ws_c.cell(row=cur_row, column=1).fill      = PatternFill("solid", fgColor="2E75B6")
                ws_c.cell(row=cur_row, column=1).alignment = s['center_align']
                ws_c.merge_cells(f"A{cur_row}:P{cur_row}")
                cur_row += 1
                for cat in df_s['Category'].unique():
                    buf_s = make_chart_sub_by_category(df_s, code, cat)
                    if buf_s is None:
                        continue
                    img_s = XLImage(buf_s)
                    img_s.width  = 720
                    img_s.height = max(220, len(df_s[df_s['Category'] == cat]) * 32 + 90)
                    ws_c.add_image(img_s, f"A{cur_row}")
                    cur_row += max(16, int(img_s.height / 15) + 2)

            cur_row += 3

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out


# =============================================
# الدالة الرئيسية
# =============================================
def run_type_3():

    set_background()
    st.markdown("<div style='margin-top: 80px;'></div>", unsafe_allow_html=True)
    st.header(" Type 3 ")

    if 'final_results_table_3' not in st.session_state:
        st.session_state.final_results_table_3 = []
    if 'detail_data_3' not in st.session_state:
        st.session_state.detail_data_3 = {}
    if 'extra_airports_3' not in st.session_state:
        st.session_state.extra_airports_3 = []

    # ── [1] ملف الأوزان ──
    st.subheader("📂 الخطوة 1: رفع ملف الأوزان ")
    weights_file = st.file_uploader(
        "ارفع ملف الأوزان (يحتوي على: Weights | subcategory weight | main category weight)",
        type=['xlsx'], key="weights_uploader_3"
    )
    if not weights_file:
        st.info("💡 ارفع ملف الأوزان أولاً للمتابعة")
        return

    try:
        xl = pd.ExcelFile(weights_file)
        required_sheets = ['Weights', 'subcategory weight', 'main category weight']
        missing = [s for s in required_sheets if s not in xl.sheet_names]
        if missing:
            st.error(f"❌ الأوراق التالية غير موجودة: {missing}")
            return

        weights_df = pd.read_excel(weights_file, sheet_name='Weights')
        weights_df.columns = weights_df.columns.str.strip()
        lookup_df = weights_df[['Ref.', 'التصنيف', 'التصنيف الفرعي', 'weight of item']].copy()
        lookup_df.columns = ['Ref.', 'التصنيف الصحيح', 'التصنيف الفرعي الصحيح', 'وزن العنصر']
        lookup_df['Ref.']         = lookup_df['Ref.'].astype(str).str.strip()
        lookup_df['_cat_norm']    = lookup_df['التصنيف الصحيح'].apply(normalize_arabic)
        lookup_df['_subcat_norm'] = lookup_df['التصنيف الفرعي الصحيح'].apply(normalize_arabic)

        sub_df = pd.read_excel(weights_file, sheet_name='subcategory weight')
        sub_df.columns = sub_df.columns.str.strip()
        year_col = 'YEAR' if 'YEAR' in sub_df.columns else 'YAER'
        display_sub = sub_df[
            (sub_df[year_col] == 2024) & (sub_df['Type'] == 3)
        ][['Category', 'Subcategory', 'Weight']].copy()
        display_sub['_cat_norm']    = display_sub['Category'].apply(normalize_arabic)
        display_sub['_subcat_norm'] = display_sub['Subcategory'].apply(normalize_arabic)

        main_df = pd.read_excel(weights_file, sheet_name='main category weight')
        main_df.columns = main_df.columns.str.strip()
        year_col_m = 'YEAR' if 'YEAR' in main_df.columns else 'YAER'
        display_main = main_df[
            (main_df[year_col_m] == 2024) & (main_df['Type'] == 3)
        ][['Category', 'Weight']].copy()
        display_main['_cat_norm'] = display_main['Category'].apply(normalize_arabic)

    except Exception as e:
        st.error(f"❌ خطأ في قراءة ملف الأوزان: {e}")
        return

    st.success("✅ تم تحميل ملف الأوزان بنجاح!")
    col1, col2 = st.columns(2)
    with col1:
        with st.expander("📊 أوزان التصنيفات الفرعية", expanded=False):
            st.dataframe(display_sub[['Category', 'Subcategory', 'Weight']], use_container_width=True)
    with col2:
        with st.expander("📊 أوزان التصنيفات الرئيسية", expanded=False):
            st.dataframe(display_main[['Category', 'Weight']], use_container_width=True)
    st.divider()

    # ── [2] ملف Airport Data ──
    st.subheader("👥 الخطوة 2: رفع ملف بيانات الركاب")
    airport_data_file = st.file_uploader(
        "ارفع ملف Airport Data — يُرفع مرة واحدة كمرجع",
        type=['xlsx', 'csv'], key="airport_data_uploader_3"
    )
    if airport_data_file:
        try:
            df_airport = (pd.read_csv(airport_data_file) if airport_data_file.name.endswith('.csv')
                          else pd.read_excel(airport_data_file))
            df_airport.columns = df_airport.columns.str.strip()
            st.session_state['airport_data_3_raw'] = df_airport
            st.session_state['airport_data_3']     = process_airport_data(df_airport.copy())
            st.success("✅ تم بنجاح تحميل  ملف الركاب:")
        except Exception as e:
            st.error(f"❌ خطأ في ملف الركاب: {e}")
            return

    if 'airport_data_3' not in st.session_state:
        st.info("💡 ارفع ملف Airport Data للمتابعة")
        return

    # بناء قواميس المطارات (الثابتة + المضافة)
    all_full_names = dict(AIRPORT_FULL_NAMES)
    all_labels     = dict(AIRPORT_LABELS)
    for ea in st.session_state.extra_airports_3:
        all_full_names[ea['code']] = ea['full_name']
        all_labels[ea['code']]     = f"🛬 {ea['code']} - {ea['label']}"

    # ملخص المطارات
    df_proc = st.session_state['airport_data_3']
    summary_rows = {}
    for code, full_name in all_full_names.items():
        row = df_proc[df_proc['Airport'] == full_name]
        if not row.empty:
            summary_rows[code] = row.iloc[0]
        elif code not in AIRPORT_FULL_NAMES:
            st.warning(
                f"⚠️ المطار **{code}** غير موجود في ملف Airport Data.\n\n"
                f"الاسم المُدخل: `{full_name}`\n\n"
                f"تأكد أن الاسم مطابق تماماً لما في الملف."
            )

    if summary_rows:
        summary_data = {'المؤشر': ['Total DOM', 'Total INT', 'Total PAX', 'DOM ratio', 'INT ratio']}
        for code, r in summary_rows.items():
            summary_data[code] = [
                f"{int(r['Total of Domestic']):,}",
                f"{int(r['Total of International']):,}",
                f"{int(r['total pas']):,}",
                f"{r['use it in for (DOM)']:.6f}",
                f"{r['use it in for (DOM)2']:.6f}",
            ]
        st.markdown("#### 📊 ملخص Airport Data")
        st.dataframe(pd.DataFrame(summary_data), use_container_width=True, hide_index=True)
    st.divider()

    # ── [3] اختيار المطار ──
    st.subheader("🏢 الخطوة 3: إدخال بيانات المطار")
    if 'selected_airport_3' not in st.session_state:
        st.session_state.selected_airport_3 = None

    # ── إضافة مطار جديد ──
    with st.expander("➕ إضافة مطار جديد (اختياري)"):
        ea1, ea2, ea3, ea4 = st.columns([1.5, 2.5, 3, 1])
        new_a_code  = ea1.text_input("كود المطار (مثال: HAE)", key="new_a_code_3").strip().upper()
        new_a_label = ea2.text_input("الاسم المختصر (مثال: حائل)", key="new_a_label_3").strip()
        # Feature 7: عرض الأسماء المتاحة من ملف الباسنجر
        available_airports = sorted(df_proc['Airport'].dropna().unique().tolist())
        new_a_full_name = ea3.selectbox(
            "اختر الاسم من ملف Airport Data",
            options=[""] + available_airports,
            key="new_a_full_3"
        )
        if ea4.button("➕ إضافة", key="add_airport_3"):
            if not new_a_code:
                st.warning("أدخل كود المطار.")
            elif not new_a_label:
                st.warning("أدخل الاسم المختصر.")
            elif not new_a_full_name:
                st.warning("أدخل الاسم الكامل.")
            elif new_a_code in all_full_names:
                st.warning(f"المطار {new_a_code} موجود مسبقاً.")
            else:
                st.session_state.extra_airports_3.append({
                    'code': new_a_code, 'label': new_a_label, 'full_name': new_a_full_name
                })
                st.rerun()

        if st.session_state.extra_airports_3:
            st.markdown("**المطارات المضافة:**")
            for ea in list(st.session_state.extra_airports_3):
                rc1, rc2 = st.columns([5, 1])
                rc1.markdown(f"- **{ea['code']}** — {ea['label']} ← `{ea['full_name']}`")
                if rc2.button("🗑️ حذف", key=f"del_airport_{ea['code']}_3"):
                    st.session_state.extra_airports_3 = [
                        x for x in st.session_state.extra_airports_3 if x['code'] != ea['code']
                    ]
                    if st.session_state.selected_airport_3 == ea['code']:
                        st.session_state.selected_airport_3 = None
                    st.rerun()

    # أزرار اختيار المطار (ديناميكية)
    st.write("✈️ اختر المطار:")
    btn_cols = st.columns(min(len(all_labels), 4))
    for i, (code, label) in enumerate(all_labels.items()):
        with btn_cols[i % 4]:
            if st.button(label, use_container_width=True,
                         type="primary" if st.session_state.selected_airport_3 == code else "secondary",
                         key=f"btn_airport_{code}_3"):
                st.session_state.selected_airport_3 = code
                st.rerun()

    if not st.session_state.selected_airport_3:
        st.info("💡 اختر المطار للمتابعة")
        return

    airport_name      = st.session_state.selected_airport_3
    airport_full_name = all_full_names[airport_name]
    st.success(f"✅ المطار المختار: {airport_name}")

    raw_file = st.file_uploader(
        f"📄 ارفع ملف Raw Data لمطار {airport_name}",
        type=['xlsx', 'csv'], key=f"raw_uploader_3_{airport_name}"
    )
    if not raw_file:
        return

    # Feature 5: مقارنة مع السنة السابقة
    with st.expander("📅 مقارنة مع السنة السابقة (اختياري)"):
        prev_raw   = st.number_input("النتيجة الخام للسنة السابقة",   min_value=0.0, max_value=1.0,
                                     value=0.0, step=0.0001, format="%.6f", key=f"prev_raw_{airport_name}")
        prev_final = st.number_input("النتيجة النهائية للسنة السابقة", min_value=0.0, max_value=1.0,
                                     value=0.0, step=0.0001, format="%.6f", key=f"prev_final_{airport_name}")
    use_prev = prev_raw > 0 or prev_final > 0

    # ── [4] الحساب ──
    with st.spinner("⏳ جاري الحساب..."):

        df_raw = (pd.read_csv(raw_file) if raw_file.name.endswith('.csv') else pd.read_excel(raw_file))
        df_raw.columns = df_raw.columns.str.strip()

        if has_star_rating(df_raw):
            st.error("❌ الملف يحتوي على (*) في عمود التقييم. يرجى تصحيحه وإعادة الرفع.")
            return

        if 'الرقم القديم' not in df_raw.columns:
            st.error("❌ لم أجد عمود 'الرقم القديم' في ملف الداتا الخام")
            return

        df_raw['Link_ID'] = df_raw['الرقم القديم'].astype(str).str.strip() + '&3'
        merged_df = pd.merge(df_raw, lookup_df, left_on='Link_ID', right_on='Ref.', how='left')
        merged_df['وزن العنصر'] = pd.to_numeric(merged_df['وزن العنصر'], errors='coerce').fillna(0)

        df_ap_raw = st.session_state['airport_data_3_raw'].copy()
        df_ap_raw.columns = df_ap_raw.columns.str.strip()
        ap_rows = df_ap_raw[df_ap_raw['Airport'] == airport_full_name].copy()
        for c in ['Domestic Arrival', 'Domestic Departure',
                  'International Arrival Total', 'International Departure Total']:
            ap_rows[c] = pd.to_numeric(ap_rows[c], errors='coerce').fillna(0)
        dom_total   = ap_rows['Domestic Arrival'].sum()   + ap_rows['Domestic Departure'].sum()
        int_total   = ap_rows['International Arrival Total'].sum() + ap_rows['International Departure Total'].sum()
        grand_total = dom_total + int_total
        dom_ratio   = dom_total / grand_total if grand_total > 0 else 1.0
        int_ratio   = int_total / grand_total if grand_total > 0 else 0.0

        J = pd.to_numeric(merged_df['التقييم الداخلي/الأساسي'], errors='coerce')
        K = pd.to_numeric(merged_df['التقييم الدولي'],          errors='coerce')
        merged_df['_J'] = J
        merged_df['_K'] = K

        def calc_item_score(row):
            j, k, w = row['_J'], row['_K'], row['وزن العنصر']
            if pd.notna(j) and pd.notna(k):
                return (((j/2) * dom_ratio) + ((k/2) * int_ratio)) * w
            elif pd.notna(j):
                return (j/2) * w
            elif pd.notna(k):
                return (k/2) * w
            return 0.0

        merged_df['Item_Score'] = merged_df.apply(calc_item_score, axis=1)

        sub_agg = merged_df.groupby(
            ['_cat_norm', '_subcat_norm', 'التصنيف الصحيح', 'التصنيف الفرعي الصحيح']
        ).agg(Item_Score_Sum=('Item_Score', 'sum'), Weight_Sum=('وزن العنصر', 'sum')).reset_index()
        sub_agg['Achievement_Calc'] = (
            sub_agg['Item_Score_Sum'] / sub_agg['Weight_Sum'].replace(0, np.nan)
        ).fillna(0)

        sub_final = pd.merge(display_sub, sub_agg, on=['_cat_norm', '_subcat_norm'], how='left')
        sub_final['Achievement_Calc'] = sub_final['Achievement_Calc'].fillna(0)
        sub_final['WEIGHT_SUM']       = sub_final['Weight'] * sub_final['Achievement_Calc']

        main_agg = sub_final.groupby(['_cat_norm', 'Category'])['WEIGHT_SUM'].sum().reset_index()
        main_agg.columns = ['_cat_norm', 'Category', 'Category_Sum_Internal']
        main_final = pd.merge(display_main, main_agg, on='_cat_norm', how='left')
        main_final.rename(columns={'Category_x': 'Category'}, inplace=True)
        main_final['Category_Sum_Internal'] = main_final['Category_Sum_Internal'].fillna(0)
        main_final['Final_Calc']            = main_final['Category_Sum_Internal'] * main_final['Weight']

        raw_score   = main_final['Final_Calc'].sum()
        final_score = raw_score / 0.85

        st.session_state.final_results_table_3.append({
            "المطار":                  airport_name,
            "النتيجة الخام":            round(raw_score, 6),
            "النتيجة النهائية (÷0.85)": round(final_score, 6)
        })

        # حفظ التفاصيل للتصدير
        disp_items = merged_df[[
            'الرقم القديم', 'التصنيف الصحيح', 'التصنيف الفرعي الصحيح',
            'التقييم الداخلي/الأساسي', 'وزن العنصر', 'Item_Score'
        ]].copy()
        disp_items.columns = ['الرقم', 'التصنيف', 'التصنيف الفرعي', 'التقييم', 'الوزن', 'Item Score']

        disp_sub_exp = sub_final[['Category', 'Subcategory', 'Weight', 'Achievement_Calc', 'WEIGHT_SUM']].copy()
        disp_sub_exp.columns = ['Category', 'Subcategory', 'الوزن', 'الإنجاز', 'الوزن × الإنجاز']

        disp_main_exp = main_final[['Category', 'Weight', 'Category_Sum_Internal', 'Final_Calc']].copy()
        disp_main_exp.columns = ['Category', 'الوزن', 'مجموع التصنيف', 'النتيجة النهائية']

        st.session_state.detail_data_3[f"Items_{airport_name}"]   = disp_items
        st.session_state.detail_data_3[f"SubCat_{airport_name}"]  = disp_sub_exp
        st.session_state.detail_data_3[f"MainCat_{airport_name}"] = disp_main_exp

    st.success(f"✅ تم الحساب بنجاح لمطار: {airport_name}")

    # Feature 5: عرض الـ delta مع السنة السابقة إذا أدخل اليوزر قيمة
    col_r1, col_r2 = st.columns(2)
    with col_r1:
        delta_raw = round(raw_score - prev_raw, 6) if use_prev else None
        st.metric("📊 النتيجة الخام", f"{raw_score*100:.4f}%",
                  delta=f"{delta_raw*100:+.4f}%" if delta_raw is not None else None)
    with col_r2:
        delta_final = round(final_score - prev_final, 6) if use_prev else None
        st.metric("🏆 النتيجة النهائية (÷ 0.85)", f"{final_score*100:.4f}%",
                  delta=f"{delta_final*100:+.4f}%" if delta_final is not None else None)

    st.divider()

    with st.expander("📋 تفاصيل العناصر - Item Level", expanded=False):
        st.dataframe(
            merged_df[['الرقم القديم', 'التصنيف الصحيح', 'التصنيف الفرعي الصحيح',
                        'التقييم الداخلي/الأساسي', 'وزن العنصر', 'Item_Score']],
            use_container_width=True)

    with st.expander("📋 تفاصيل التصنيف الفرعي - Subcategory Level", expanded=True):
        df_sub_display = sub_final[['Category', 'Subcategory', 'Weight', 'Achievement_Calc', 'WEIGHT_SUM']].copy()
        df_sub_display['Achievement_Calc'] = df_sub_display['Achievement_Calc'] * 100
        df_sub_display['WEIGHT_SUM']       = df_sub_display['WEIGHT_SUM'] * 100
        df_sub_display.columns = ['Category', 'Subcategory', 'Weight (Sub)', 'Achievement %', 'WEIGHT × Achievement %']
        st.dataframe(df_sub_display, use_container_width=True,
            column_config={
                'Weight (Sub)':             st.column_config.NumberColumn(format="%.6f"),
                'Achievement %':            st.column_config.NumberColumn(format="%.4f%%"),
                'WEIGHT × Achievement %':   st.column_config.NumberColumn(format="%.4f%%"),
            })

    with st.expander("📋 تفاصيل التصنيف الرئيسي - Main Category Level", expanded=True):
        df_main_display = main_final[['Category', 'Weight', 'Category_Sum_Internal', 'Final_Calc']].copy()
        df_main_display['Category_Sum_Internal'] = df_main_display['Category_Sum_Internal'] * 100
        df_main_display['Final_Calc']            = df_main_display['Final_Calc'] * 100
        df_main_display.columns = ['Category', 'Weight (Main)', 'مجموع التصنيف %', 'النتيجة النهائية %']
        st.dataframe(df_main_display, use_container_width=True,
            column_config={
                'مجموع التصنيف %':   st.column_config.NumberColumn(format="%.4f%%"),
                'Weight (Main)':     st.column_config.NumberColumn(format="%.6f"),
                'النتيجة النهائية %': st.column_config.NumberColumn(format="%.4f%%"),
            })

    unmatched = merged_df[merged_df['التصنيف الصحيح'].isna()]
    if len(unmatched) > 0:
        with st.expander(f"⚠️ عناصر غير مرتبطة ({len(unmatched)} عنصر)", expanded=False):
            st.warning("هذه العناصر لم تُطابَق مع ملف الأوزان")
            st.dataframe(
                unmatched[['الرقم القديم', 'Link_ID', 'التقييم الداخلي/الأساسي']],
                use_container_width=True)

    # ── الجدول التراكمي + التحميل ──
    if st.session_state.final_results_table_3:
        st.divider()
        st.subheader("🏆 الجدول النهائي لجميع المطارات")

        df_res = pd.DataFrame(st.session_state.final_results_table_3).copy()
        df_res['النتيجة الخام']            = df_res['النتيجة الخام'] * 100
        df_res['النتيجة النهائية (÷0.85)'] = df_res['النتيجة النهائية (÷0.85)'] * 100
        st.dataframe(df_res, use_container_width=True,
            column_config={
                'النتيجة الخام':            st.column_config.NumberColumn(format="%.4f%%"),
                'النتيجة النهائية (÷0.85)': st.column_config.NumberColumn(format="%.4f%%"),
            })

        # ── خيارات التحميل ──
        st.divider()
        st.subheader("⬇️ تحميل النتائج")
        dl_col1, dl_col2, dl_col3 = st.columns(3)

        with dl_col1:
            st.markdown("**📊 Excel كامل**")
            st.caption("النتائج التراكمية + تفاصيل Items + SubCategory + Main Category")
            excel_full = build_excel_report(
                st.session_state.final_results_table_3,
                detail_data=st.session_state.detail_data_3
            )
            st.download_button(
                label="📥 تحميل Excel الكامل",
                data=excel_full,
                file_name="Type3_Full_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_full_excel_3"
            )

        with dl_col2:
            st.markdown("**📋 Excel تراكمي فقط**")
            st.caption("الجدول التراكمي النهائي فقط بدون التفاصيل")
            excel_summary = build_excel_report(
                st.session_state.final_results_table_3,
                detail_data=None
            )
            st.download_button(
                label="📥 تحميل Excel التراكمي",
                data=excel_summary,
                file_name="Type3_Summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_summary_excel_3"
            )

        with dl_col3:
            st.markdown("**📈 Excel مع رسومات**")
            st.caption("النتائج + التفاصيل + Bar Charts للتصنيفات الرئيسية والفرعية")
            if st.session_state.detail_data_3:
                with st.spinner("جاري إنشاء الرسومات..."):
                    excel_charts = build_excel_with_charts(
                        st.session_state.final_results_table_3,
                        st.session_state.detail_data_3
                    )
                st.download_button(
                    label="📥 تحميل Excel + رسومات",
                    data=excel_charts,
                    file_name="Type3_Charts_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_charts_excel_3"
                )
            else:
                st.info("احسب مطاراً أولاً لتفعيل هذا الخيار")

        st.divider()
        if st.button("🗑️ مسح جميع النتائج"):
            st.session_state.final_results_table_3 = []
            st.session_state.detail_data_3 = {}
            st.rerun()
