import streamlit as st
import pandas as pd
import numpy as np
import re
import base64
import os
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# =============================================
# دالة تطبيع النص العربي
# =============================================
def normalize_arabic(text):
    if not isinstance(text, str):
        return text
    text = text.strip()
    text = re.sub(r'[أإآا]', 'ا', text)
    text = re.sub(r'[ىي]', 'ي', text)
    text = re.sub(r'[ةه]', 'ه', text)
    text = re.sub(r'[\u064B-\u065F]', '', text)
    return text


# =============================================
# دالة تحميل الخلفية
# =============================================
def set_background():
    for ext in ['png', 'jpg', 'jpeg', 'webp']:
        path = f"Background.{ext}"
        if os.path.exists(path):
            with open(path, "rb") as f:
                data = base64.b64encode(f.read()).decode()
            st.markdown(f"""
                <style>
                .stApp {{
                    background-image: url("data:image/{ext};base64,{data}");
                    background-size: cover;
                    background-position: center;
                    background-attachment: fixed;
                }}
                .block-container {{
                    background-color: rgba(255, 255, 255, 0.82);
                    border-radius: 12px;
                    padding: 1.5rem;
                }}
                .stButton > button[type="primary"] {{
                    background-color: #003366 !important;
                    border-color: #003366 !important;
                    color: white !important;
                }}
                </style>
            """, unsafe_allow_html=True)
            return


# =============================================
# دالة حساب DOM/INT ratio من ملف الباسنجر
# =============================================
def get_dom_int_ratios(df_pas, terminal_code):
    for col in ['Domestic Arrival', 'Domestic Departure',
                'International Arrival Total', 'International Departure Total']:
        if col in df_pas.columns:
            df_pas[col] = pd.to_numeric(df_pas[col], errors='coerce').fillna(0)

    rows = pd.DataFrame()
    if 'Airport Terminal' in df_pas.columns:
        rows = df_pas[df_pas['Airport Terminal'] == terminal_code]

    if not rows.empty:
        dom   = rows['Domestic Arrival'].sum() + rows['Domestic Departure'].sum()
        intl  = rows['International Arrival Total'].sum() + rows['International Departure Total'].sum()
        grand = dom + intl
        if grand == 0:
            return 1.0, 0.0, 0, 0, 0
        return dom / grand, intl / grand, int(dom), int(intl), int(grand)

    if 'Airport' in df_pas.columns:
        airport_name = terminal_code.split('>')[0].strip()
        rows2 = df_pas[df_pas['Airport'] == airport_name]
        if not rows2.empty:
            dom   = rows2['Domestic Arrival'].sum() + rows2['Domestic Departure'].sum()
            intl  = rows2['International Arrival Total'].sum() + rows2['International Departure Total'].sum()
            grand = dom + intl
            if grand > 0:
                return dom / grand, intl / grand, int(dom), int(intl), int(grand)

    if 'Airport.1' in df_pas.columns and 'use it in for (DOM)' in df_pas.columns:
        keyword = terminal_code.split('>')[0].strip()[:20]
        rows2 = df_pas[df_pas['Airport.1'].astype(str).str.contains(keyword, case=False, na=False)]
        if not rows2.empty:
            dom_r = pd.to_numeric(rows2['use it in for (DOM)'].iloc[0],  errors='coerce')
            int_r = pd.to_numeric(rows2['use it in for (DOM)2'].iloc[0], errors='coerce')
            total = pd.to_numeric(rows2['total pas'].iloc[0],             errors='coerce')
            if pd.notna(dom_r) and pd.notna(int_r):
                dom  = int(dom_r * total) if pd.notna(total) else 0
                intl = int(int_r * total) if pd.notna(total) else 0
                return float(dom_r), float(int_r), dom, intl, int(total) if pd.notna(total) else 0

    return 1.0, 0.0, 0, 0, 0


# =============================================
# التحقق من وجود * في ملف المطار
# =============================================
def has_star_rating(df):
    col = 'التقييم الداخلي/الأساسي'
    if col not in df.columns:
        return False
    return df[col].astype(str).str.strip().eq('*').any()


# =============================================
# حساب Item Score
# =============================================
def calc_item_score(j, k, dom_r, int_r):
    if pd.notna(j) and pd.notna(k):
        return (j / 2 * dom_r) + (k / 2 * int_r)
    elif pd.notna(j):
        return j / 2
    elif pd.notna(k):
        return k / 2
    return np.nan


# =============================================
# دوال Excel
# =============================================
def _excel_styles():
    return dict(
        header_fill  = PatternFill("solid", fgColor="1F4E79"),
        alt_fill     = PatternFill("solid", fgColor="D6E4F0"),
        white_fill   = PatternFill("solid", fgColor="FFFFFF"),
        green_fill   = PatternFill("solid", fgColor="E2EFDA"),
        bold_white   = Font(name="Arial", bold=True, color="FFFFFF", size=11),
        normal_font  = Font(name="Arial", size=10),
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True),
        thin_border  = Border(
            left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin',  color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
        )
    )

def _sh(ws, row, cols, s):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = s['header_fill']; cell.font = s['bold_white']
        cell.alignment = s['center_align']; cell.border = s['thin_border']

def _sd(ws, row, cols, s, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = s['alt_fill'] if alt else s['white_fill']
        cell.font = s['normal_font']
        cell.alignment = s['center_align']; cell.border = s['thin_border']

def _write_summary_sheet(wb, results_table, s):
    ws1 = wb.active
    ws1.title = "النتائج التراكمية"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "الجدول التراكمي النهائي - Type 2"
    ws1["A1"].font      = Font(name="Arial", bold=True, color="1F4E79", size=14)
    ws1["A1"].alignment = s['center_align']
    ws1["A1"].fill      = PatternFill("solid", fgColor="D6E4F0")
    ws1.row_dimensions[1].height = 30
    for i, h in enumerate(["المطار", "النتيجة الخام", "النتيجة النهائية (÷0.85)"], 1):
        ws1.cell(row=2, column=i, value=h)
    _sh(ws1, 2, 3, s)
    for r_idx, row in enumerate(results_table, start=3):
        ws1.cell(row=r_idx, column=1, value=row["المطار"])
        ws1.cell(row=r_idx, column=2, value=row["النتيجة الخام"])
        ws1.cell(row=r_idx, column=3, value=row["النتيجة النهائية (÷0.85)"])
        _sd(ws1, r_idx, 3, s, alt=(r_idx % 2 == 0))
        ws1.cell(row=r_idx, column=2).number_format = "0.0000%"
        ws1.cell(row=r_idx, column=3).number_format = "0.0000%"
        ws1.cell(row=r_idx, column=3).fill = s['green_fill']
        ws1.cell(row=r_idx, column=3).font = Font(name="Arial", bold=True, color="375623", size=10)
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 28

def _write_detail_sheets(wb, detail_data, s):
    for sheet_name, df_detail in detail_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.sheet_view.rightToLeft = True
        for c_idx, col_name in enumerate(df_detail.columns, 1):
            ws.cell(row=1, column=c_idx, value=col_name)
        _sh(ws, 1, len(df_detail.columns), s)
        for r_idx, row in enumerate(df_detail.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            _sd(ws, r_idx, len(df_detail.columns), s, alt=(r_idx % 2 == 0))
        for c_idx, col_name in enumerate(df_detail.columns, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = max(18, len(str(col_name)) + 4)

def build_excel_report(results_table, detail_data=None):
    s = _excel_styles(); wb = Workbook()
    _write_summary_sheet(wb, results_table, s)
    if detail_data: _write_detail_sheets(wb, detail_data, s)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

def fix_arabic(text):
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        return get_display(arabic_reshaper.reshape(str(text)))
    except Exception:
        return str(text)

def make_chart_main(df_main, airport_code):
    categories = [fix_arabic(c) for c in df_main['Category'].tolist()]
    values     = [float(v) * 100 for v in df_main['النتيجة النهائية'].tolist()]
    fig, ax = plt.subplots(figsize=(13, max(5, len(categories) * 0.55 + 2)))
    fig.patch.set_facecolor('#F7FBFF'); ax.set_facecolor('#F7FBFF')
    colors = ['#1F4E79' if v >= 70 else '#2E75B6' if v >= 50 else '#9DC3E6' for v in values]
    bars = ax.barh(categories, values, color=colors, edgecolor='white', height=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.2, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va='center', ha='left', fontsize=9, color='#1F4E79', fontweight='bold')
    ax.set_xlim(0, max(values) * 1.3 if values else 100)
    ax.set_title(fix_arabic(f'نتائج التصنيفات الرئيسية - {airport_code}'),
                 fontsize=13, fontweight='bold', color='#1F4E79', pad=15)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.4, color='#BFBFBF')
    from matplotlib.patches import Patch
    ax.legend(handles=[
        Patch(facecolor='#1F4E79', label=fix_arabic('ممتاز (≥ 70%)')),
        Patch(facecolor='#2E75B6', label=fix_arabic('جيد (≥ 50%)')),
        Patch(facecolor='#9DC3E6', label=fix_arabic('يحتاج تحسين (< 50%)')),
    ], loc='lower right', fontsize=8, framealpha=0.7)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); buf.seek(0)
    return buf

def make_chart_sub(df_sub, airport_code, category_name):
    df_cat = df_sub[df_sub['Category'] == category_name].copy()
    if df_cat.empty: return None
    labels = [fix_arabic(r['Subcategory']) for _, r in df_cat.iterrows()]
    values = [float(v) * 100 for v in df_cat['الوزن × الإنجاز'].tolist()]
    fig, ax = plt.subplots(figsize=(12, max(3.5, len(labels) * 0.5 + 1.5)))
    fig.patch.set_facecolor('#F7FBFF'); ax.set_facecolor('#F7FBFF')
    palette = plt.cm.Blues(np.linspace(0.45, 0.88, max(len(labels), 1)))
    bars = ax.barh(labels, values, color=palette, edgecolor='white', height=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va='center', ha='left', fontsize=9, color='#1F4E79')
    ax.set_xlim(0, max(values) * 1.3 if values else 100)
    ax.set_title(fix_arabic(f'{category_name} | {airport_code}'),
                 fontsize=11, fontweight='bold', color='#1F4E79', pad=10)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.4, color='#BFBFBF')
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); buf.seek(0)
    return buf

def build_excel_with_charts(results_table, detail_data):
    s = _excel_styles(); wb = Workbook()
    _write_summary_sheet(wb, results_table, s)
    if detail_data: _write_detail_sheets(wb, detail_data, s)
    airports_in_data = {k.replace("MainCat_", "") for k in detail_data if k.startswith("MainCat_")}
    if airports_in_data:
        ws_c = wb.create_sheet(title="الرسومات البيانية")
        ws_c.sheet_view.rightToLeft = True
        ws_c["A1"] = "الرسومات البيانية"
        ws_c["A1"].font = Font(name="Arial", bold=True, color="1F4E79", size=14)
        ws_c["A1"].fill = PatternFill("solid", fgColor="D6E4F0")
        ws_c["A1"].alignment = s['center_align']
        cur_row = 3
        for code in sorted(airports_in_data):
            ws_c.cell(row=cur_row, column=1, value=f"مطار {code}")
            ws_c.cell(row=cur_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
            ws_c.cell(row=cur_row, column=1).fill = PatternFill("solid", fgColor="1F4E79")
            ws_c.cell(row=cur_row, column=1).alignment = s['center_align']
            ws_c.merge_cells(f"A{cur_row}:P{cur_row}"); cur_row += 1
            if f"MainCat_{code}" in detail_data:
                df_m = detail_data[f"MainCat_{code}"]
                buf_m = make_chart_main(df_m, code)
                img_m = XLImage(buf_m)
                img_m.width = 750; img_m.height = max(320, len(df_m) * 28 + 80)
                ws_c.add_image(img_m, f"A{cur_row}")
                cur_row += max(22, int(img_m.height / 15) + 2)
            if f"SubCat_{code}" in detail_data:
                df_s = detail_data[f"SubCat_{code}"]
                ws_c.cell(row=cur_row, column=1, value=f"التصنيفات الفرعية - {code}")
                ws_c.cell(row=cur_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                ws_c.cell(row=cur_row, column=1).fill = PatternFill("solid", fgColor="2E75B6")
                ws_c.cell(row=cur_row, column=1).alignment = s['center_align']
                ws_c.merge_cells(f"A{cur_row}:P{cur_row}"); cur_row += 1
                for cat in df_s['Category'].unique():
                    buf_s = make_chart_sub(df_s, code, cat)
                    if buf_s is None: continue
                    img_s = XLImage(buf_s)
                    img_s.width = 720; img_s.height = max(220, len(df_s[df_s['Category'] == cat]) * 32 + 90)
                    ws_c.add_image(img_s, f"A{cur_row}")
                    cur_row += max(16, int(img_s.height / 15) + 2)
            cur_row += 3
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out


# =============================================
# الدالة الرئيسية
# =============================================
def run_type_2():

    set_background()
    st.markdown("<div style='margin-top: 80px;'></div>", unsafe_allow_html=True)
    st.header("📊 معالجة Type 2 - الدمام والمدينة")

    if 'final_results_table_2' not in st.session_state:
        st.session_state.final_results_table_2 = []
    if 'detail_data_2' not in st.session_state:
        st.session_state.detail_data_2 = {}
    if 'extra_airports_2' not in st.session_state:
        st.session_state.extra_airports_2 = []

    # [1] ملف الأوزان
    st.subheader("📂 الخطوة 1: رفع ملف الأوزان")
    weights_file = st.file_uploader(
        "ارفع ملف الأوزان", type=['xlsx'], key="weights_uploader_2"
    )
    if not weights_file:
        st.info("💡 ارفع ملف الأوزان أولاً للمتابعة")
        return

    try:
        w_items = pd.read_excel(weights_file, sheet_name='Weights')
        w_items.columns = w_items.columns.str.strip()
        w2 = w_items[w_items['Type'] == 2].copy()

        lookup_df = w2[['Ref.', 'التصنيف', 'التصنيف الفرعي', 'weight of item']].copy()
        lookup_df.columns = ['Ref.', 'التصنيف الصحيح', 'التصنيف الفرعي الصحيح', 'وزن العنصر']
        lookup_df['Ref.']         = lookup_df['Ref.'].astype(str).str.strip()
        lookup_df['_cat_norm']    = lookup_df['التصنيف الصحيح'].apply(normalize_arabic)
        lookup_df['_subcat_norm'] = lookup_df['التصنيف الفرعي الصحيح'].apply(normalize_arabic)

        sub_raw = pd.read_excel(weights_file, sheet_name='subcategory weight')
        sub_raw.columns = sub_raw.columns.str.strip()
        year_col = 'YEAR' if 'YEAR' in sub_raw.columns else 'YAER'
        display_sub = sub_raw[
            (sub_raw['Type'] == 2) & (sub_raw[year_col] == 2024)
        ][['Category', 'Subcategory', 'Weight']].copy()
        display_sub['_cat_norm']    = display_sub['Category'].apply(normalize_arabic)
        display_sub['_subcat_norm'] = display_sub['Subcategory'].apply(normalize_arabic)

        main_raw = pd.read_excel(weights_file, sheet_name='main category weight')
        main_raw.columns = main_raw.columns.str.strip()
        year_col_m = 'YEAR' if 'YEAR' in main_raw.columns else 'YAER'
        display_main = main_raw[
            (main_raw['Type'] == 2) & (main_raw[year_col_m] == 2024)
        ][['Category', 'Weight']].copy()
        display_main['_cat_norm'] = display_main['Category'].apply(normalize_arabic)

    except Exception as e:
        st.error(f"❌ خطأ في ملف الأوزان: {e}")
        return

    st.success("✅ تم تحميل ملف الأوزان")
    cw1, cw2 = st.columns(2)
    with cw1:
        with st.expander("📊 أوزان التصنيفات الفرعية", expanded=False):
            st.dataframe(display_sub[['Category', 'Subcategory', 'Weight']], use_container_width=True)
    with cw2:
        with st.expander("📊 أوزان التصنيفات الرئيسية", expanded=False):
            st.dataframe(display_main[['Category', 'Weight']], use_container_width=True)

    st.divider()

    # [2] ملف الباسنجر
    st.subheader("👥 الخطوة 2: رفع ملف بيانات الباسنجر")
    pas_file = st.file_uploader(
        "ارفع ملف Airport Data", type=['xlsx', 'csv'], key="pas_uploader_2"
    )
    if not pas_file:
        st.info("💡 ارفع ملف الباسنجر للمتابعة")
        return

    try:
        df_pas = (pd.read_csv(pas_file) if pas_file.name.endswith('.csv')
                  else pd.read_excel(pas_file))
        df_pas.columns = df_pas.columns.str.strip()
    except Exception as e:
        st.error(f"❌ خطأ في ملف الباسنجر: {e}")
        return

    # بناء قائمة المطارات (الثابتة + المضافة)
    AIRPORTS = {
        'DMM': {'label': 'الدمام (DMM)', 'terminal': 'King Fahd International Airport>1'},
        'MED': {'label': 'المدينة (MED)', 'terminal': 'Prince Mohammed bin Abdulaziz International Airport>1'},
    }
    for ea in st.session_state.extra_airports_2:
        AIRPORTS[ea['code']] = {'label': ea['label'], 'terminal': ea['terminal']}

    pax_data = {}
    for code, info in AIRPORTS.items():
        dom_r, int_r, dom, intl, total = get_dom_int_ratios(df_pas.copy(), info['terminal'])
        pax_data[code] = {'dom_r': dom_r, 'int_r': int_r, 'dom': dom, 'intl': intl, 'total': total}

    st.success("✅ تم تحميل ملف الباسنجر")

    summary_rows = {'المؤشر': ['DOM', 'INT', 'الإجمالي', 'نسبة DOM', 'نسبة INT']}
    for code, info in AIRPORTS.items():
        summary_rows[info['label']] = [
            f"{pax_data[code]['dom']:,}", f"{pax_data[code]['intl']:,}",
            f"{pax_data[code]['total']:,}", f"{pax_data[code]['dom_r']:.4f}", f"{pax_data[code]['int_r']:.4f}",
        ]
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    st.divider()

    # [3] اختيار المطار ورفع الملف الخام
    st.subheader("🏢 الخطوة 3: اختر المطار وارفع ملف البيانات")
    st.caption("⚠️ الملفات التي تحتوي على (*) في عمود التقييم لن تُقبل")

    if 'selected_airport_2' not in st.session_state:
        st.session_state.selected_airport_2 = None

    # ── إضافة مطار جديد ──
    with st.expander("➕ إضافة مطار جديد (اختياري)"):
        ea1, ea2, ea3, ea4 = st.columns([1.5, 2.5, 3, 1])
        new_a_code     = ea1.text_input("كود المطار (مثال: JED)", key="new_a_code_2").strip().upper()
        new_a_label    = ea2.text_input("الاسم (مثال: جدة (JED))", key="new_a_label_2").strip()
        available_terminals_2 = sorted(df_pas['Airport Terminal'].dropna().unique().tolist()) if 'Airport Terminal' in df_pas.columns else []
        new_a_terminal = ea3.selectbox("اختر الصالة من ملف الباسنجر", options=[""] + available_terminals_2, key="new_a_terminal_2")
        if ea4.button("➕ إضافة", key="add_airport_2"):
            existing_codes = list(AIRPORTS.keys())
            if not new_a_code:
                st.warning("أدخل كود المطار.")
            elif not new_a_label:
                st.warning("أدخل اسم المطار.")
            elif not new_a_terminal:
                st.warning("أدخل كود الصالة.")
            elif new_a_code in existing_codes:
                st.warning(f"المطار {new_a_code} موجود مسبقاً.")
            else:
                st.session_state.extra_airports_2.append({
                    'code': new_a_code, 'label': new_a_label, 'terminal': new_a_terminal
                })
                st.rerun()

        if st.session_state.extra_airports_2:
            st.markdown("**المطارات المضافة:**")
            for ea in list(st.session_state.extra_airports_2):
                rc1, rc2 = st.columns([5, 1])
                rc1.markdown(f"- **{ea['code']}** — {ea['label']} ← `{ea['terminal']}`")
                if rc2.button("🗑️ حذف", key=f"del_airport_{ea['code']}_2"):
                    st.session_state.extra_airports_2 = [
                        x for x in st.session_state.extra_airports_2 if x['code'] != ea['code']
                    ]
                    if st.session_state.selected_airport_2 == ea['code']:
                        st.session_state.selected_airport_2 = None
                    st.rerun()

    # أزرار اختيار المطار (ديناميكية)
    btn_cols = st.columns(min(len(AIRPORTS), 4))
    for i, (code, info) in enumerate(AIRPORTS.items()):
        with btn_cols[i % 4]:
            if st.button(f"🛬 {code} - {info['label']}", use_container_width=True,
                         type="primary" if st.session_state.selected_airport_2 == code else "secondary",
                         key=f"btn_airport_{code}_2"):
                st.session_state.selected_airport_2 = code
                st.rerun()

    if not st.session_state.selected_airport_2:
        st.info("💡 اختر المطار للمتابعة")
        return

    airport_code = st.session_state.selected_airport_2
    st.success(f"✅ المطار المختار: {AIRPORTS[airport_code]['label']}")

    dom_r = pax_data[airport_code]['dom_r']
    int_r = pax_data[airport_code]['int_r']

    raw_file = st.file_uploader(
        f"📄 ارفع ملف Raw Data لمطار {airport_code}",
        type=['xlsx'],
        key=f"raw_uploader_2_{airport_code}"
    )
    if not raw_file:
        return

    # ── مقارنة مع السنة السابقة ──
    with st.expander("📅 مقارنة مع السنة السابقة (اختياري)"):
        py1, py2 = st.columns(2)
        prev_raw_2   = py1.number_input("النتيجة الخام للسنة السابقة",   min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key=f"prev_raw_2_{airport_code}")
        prev_final_2 = py2.number_input("النتيجة النهائية للسنة السابقة", min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key=f"prev_final_2_{airport_code}")
    use_prev_2 = prev_raw_2 > 0 or prev_final_2 > 0

    # [4] الحساب
    if st.button("🚀 ابدأ الحساب", type="primary", key=f"calc_2_{airport_code}"):

        with st.spinner("⏳ جاري الحساب..."):
            try:
                df_raw = pd.read_excel(raw_file)
                df_raw.columns = df_raw.columns.str.strip()

                if has_star_rating(df_raw):
                    st.error("❌ الملف يحتوي على (*) في عمود التقييم. يرجى تصحيحه وإعادة الرفع.")
                    return

                df_raw['الرقم القديم'] = pd.to_numeric(df_raw['الرقم القديم'], errors='coerce')
                df_raw['Link_ID'] = df_raw['الرقم القديم'].apply(
                    lambda x: f"{int(x)}&2" if pd.notna(x) else ''
                )

                merged = pd.merge(df_raw, lookup_df, left_on='Link_ID', right_on='Ref.', how='inner')
                merged['وزن العنصر'] = pd.to_numeric(merged['وزن العنصر'], errors='coerce').fillna(0)

                J = pd.to_numeric(merged['التقييم الداخلي/الأساسي'], errors='coerce')
                K = pd.to_numeric(merged['التقييم الدولي'] if 'التقييم الدولي' in merged.columns else pd.Series(dtype=float), errors='coerce')

                merged['item_score'] = [
                    calc_item_score(j, k, dom_r, int_r)
                    for j, k in zip(J, K)
                ]
                merged['Item_Score_x_Weight'] = merged['item_score'].fillna(0) * merged['وزن العنصر']

                sub_agg = merged.groupby(
                    ['_cat_norm', '_subcat_norm', 'التصنيف الصحيح', 'التصنيف الفرعي الصحيح']
                ).agg(
                    Item_Score_Sum=('Item_Score_x_Weight', 'sum'),
                    Weight_Sum=('وزن العنصر', 'sum')
                ).reset_index()
                sub_agg['Achievement'] = (
                    sub_agg['Item_Score_Sum'] / sub_agg['Weight_Sum'].replace(0, np.nan)
                ).fillna(0)

                sub_final = pd.merge(
                    display_sub,
                    sub_agg[['_cat_norm', '_subcat_norm', 'Achievement']],
                    on=['_cat_norm', '_subcat_norm'], how='left'
                )
                sub_final['Achievement'] = sub_final['Achievement'].fillna(0)
                sub_final['WEIGHT_SUM']  = sub_final['Weight'] * sub_final['Achievement']

                main_agg = sub_final.groupby(
                    ['_cat_norm', 'Category']
                )['WEIGHT_SUM'].sum().reset_index()
                main_agg.columns = ['_cat_norm', 'Category', 'Cat_Sum']

                main_final = pd.merge(display_main, main_agg, on='_cat_norm', how='left')
                main_final.rename(columns={'Category_x': 'Category'}, inplace=True)
                main_final['Cat_Sum']    = main_final['Cat_Sum'].fillna(0)
                main_final['Final_Calc'] = main_final['Cat_Sum'] * main_final['Weight']

                raw_score   = main_final['Final_Calc'].sum()
                final_score = raw_score / 0.85

                st.session_state.final_results_table_2.append({
                    "المطار":                  airport_code,
                    "النتيجة الخام":            round(raw_score, 6),
                    "النتيجة النهائية (÷0.85)": round(final_score, 6),
                })

                # حفظ التفاصيل
                disp_items = merged[['الرقم القديم','التصنيف الصحيح','التصنيف الفرعي الصحيح','التقييم الداخلي/الأساسي','التقييم الدولي','item_score','وزن العنصر','Item_Score_x_Weight']].copy()
                disp_items.columns = ['الرقم','التصنيف','التصنيف الفرعي','J (داخلي)','K (دولي)','Item Score','الوزن','Score × الوزن']
                disp_sub_exp = sub_final[['Category','Subcategory','Weight','Achievement','WEIGHT_SUM']].copy()
                disp_sub_exp.columns = ['Category','Subcategory','الوزن','الإنجاز','الوزن × الإنجاز']
                disp_main_exp = main_final[['Category','Weight','Cat_Sum','Final_Calc']].copy()
                disp_main_exp.columns = ['Category','الوزن','مجموع التصنيف','النتيجة النهائية']

                st.session_state.detail_data_2[f"Items_{airport_code}"]   = disp_items
                st.session_state.detail_data_2[f"SubCat_{airport_code}"]  = disp_sub_exp
                st.session_state.detail_data_2[f"MainCat_{airport_code}"] = disp_main_exp

            except Exception as e:
                st.error(f"❌ خطأ في الحساب: {e}")
                import traceback
                st.code(traceback.format_exc())
                return

        # [5] عرض النتائج
        st.success(f"✅ تم الحساب بنجاح لمطار {airport_code}!")

        cr1, cr2 = st.columns(2)
        _delta_raw_2   = f"{(raw_score - prev_raw_2)*100:.4f}%"   if use_prev_2 else None
        _delta_final_2 = f"{(final_score - prev_final_2)*100:.4f}%" if use_prev_2 else None
        cr1.metric("📊 النتيجة الخام",            f"{raw_score*100:.4f}%",   delta=_delta_raw_2)
        cr2.metric("🏆 النتيجة النهائية (÷ 0.85)", f"{final_score*100:.4f}%", delta=_delta_final_2)

        st.divider()

        with st.expander("📋 تفاصيل العناصر - Item Level", expanded=False):
            disp = merged[['الرقم القديم','التصنيف الصحيح','التصنيف الفرعي الصحيح','التقييم الداخلي/الأساسي','التقييم الدولي','item_score','وزن العنصر','Item_Score_x_Weight']].copy()
            disp.columns = ['الرقم','التصنيف','التصنيف الفرعي','J (داخلي)','K (دولي)','Item Score','الوزن','Score × الوزن']
            st.dataframe(disp, use_container_width=True,
                column_config={
                    'Item Score':    st.column_config.NumberColumn(format="%.4f"),
                    'الوزن':         st.column_config.NumberColumn(format="%.6f"),
                    'Score × الوزن': st.column_config.NumberColumn(format="%.6f"),
                })

        with st.expander("📋 تفاصيل التصنيف الفرعي", expanded=True):
            disp_sub = sub_final[['Category','Subcategory','Weight','Achievement','WEIGHT_SUM']].copy()
            disp_sub.columns = ['Category','Subcategory','الوزن','الإنجاز','الوزن × الإنجاز']
            disp_sub['الإنجاز']         = disp_sub['الإنجاز'] * 100
            disp_sub['الوزن × الإنجاز'] = disp_sub['الوزن × الإنجاز'] * 100
            st.dataframe(disp_sub, use_container_width=True,
                column_config={
                    'الوزن':           st.column_config.NumberColumn(format="%.6f"),
                    'الإنجاز':         st.column_config.NumberColumn(format="%.4f%%"),
                    'الوزن × الإنجاز': st.column_config.NumberColumn(format="%.4f%%"),
                })

        with st.expander("📋 تفاصيل التصنيف الرئيسي", expanded=True):
            disp_main = main_final[['Category','Weight','Cat_Sum','Final_Calc']].copy()
            disp_main.columns = ['Category','الوزن','مجموع التصنيف','النتيجة النهائية']
            disp_main['مجموع التصنيف']    = disp_main['مجموع التصنيف'] * 100
            disp_main['النتيجة النهائية'] = disp_main['النتيجة النهائية'] * 100
            st.dataframe(disp_main, use_container_width=True,
                column_config={
                    'الوزن':            st.column_config.NumberColumn(format="%.6f"),
                    'مجموع التصنيف':    st.column_config.NumberColumn(format="%.4f%%"),
                    'النتيجة النهائية': st.column_config.NumberColumn(format="%.4f%%"),
                })

        unmatched = merged[merged['التصنيف الصحيح'].isna()]
        if len(unmatched) > 0:
            with st.expander(f"⚠️ عناصر غير مرتبطة ({len(unmatched)} عنصر)", expanded=False):
                st.warning("هذه العناصر لم تُطابَق مع ملف الأوزان")
                st.dataframe(unmatched[['الرقم القديم','Link_ID']], use_container_width=True)

    # الجدول التراكمي النهائي
    if st.session_state.final_results_table_2:
        st.divider()
        st.subheader("🏆 الجدول النهائي التراكمي")
        df_res_2 = pd.DataFrame(st.session_state.final_results_table_2).copy()
        df_res_2['النتيجة الخام']            = df_res_2['النتيجة الخام'] * 100
        df_res_2['النتيجة النهائية (÷0.85)'] = df_res_2['النتيجة النهائية (÷0.85)'] * 100
        st.dataframe(df_res_2, use_container_width=True,
            column_config={
                'النتيجة الخام':            st.column_config.NumberColumn(format="%.4f%%"),
                'النتيجة النهائية (÷0.85)': st.column_config.NumberColumn(format="%.4f%%"),
            })

        # خيارات التحميل
        st.divider()
        st.subheader("⬇️ تحميل النتائج")
        dl1, dl2, dl3 = st.columns(3)

        with dl1:
            st.markdown("**📊 Excel كامل**")
            st.caption("النتائج + تفاصيل Items + SubCategory + Main Category")
            st.download_button("📥 تحميل Excel الكامل",
                data=build_excel_report(st.session_state.final_results_table_2, st.session_state.detail_data_2),
                file_name="Type2_Full_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_full_2")

        with dl2:
            st.markdown("**📋 Excel تراكمي فقط**")
            st.caption("الجدول التراكمي فقط")
            st.download_button("📥 تحميل Excel التراكمي",
                data=build_excel_report(st.session_state.final_results_table_2),
                file_name="Type2_Summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_summary_2")

        with dl3:
            st.markdown("**📈 Excel مع رسومات**")
            st.caption("النتائج + Bar Charts")
            if st.session_state.detail_data_2:
                with st.spinner("جاري إنشاء الرسومات..."):
                    excel_charts = build_excel_with_charts(
                        st.session_state.final_results_table_2, st.session_state.detail_data_2)
                st.download_button("📥 تحميل Excel + رسومات",
                    data=excel_charts, file_name="Type2_Charts_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dl_charts_2")

        st.divider()
        if st.button("🗑️ مسح النتائج", key="clear_2"):
            st.session_state.final_results_table_2 = []
            st.session_state.detail_data_2 = {}
            st.rerun()


if __name__ == "__main__":
    run_type_2()