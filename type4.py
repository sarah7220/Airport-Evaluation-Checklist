import streamlit as st
import pandas as pd
import numpy as np
import re
import base64
import os
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# 
# دالة  النص العربي
# 
def normalize_arabic(text):
    if not isinstance(text, str):
        return text
    text = text.strip()
    text = re.sub(r'[أإآا]', 'ا', text)
    text = re.sub(r'[ىي]', 'ي', text)
    text = re.sub(r'[ةه]', 'ه', text)
    text = re.sub(r'[\u064B-\u065F]', '', text)
    return text

# 
# دالة تحميل الخلفية (نفس الديزاين والألوان)
# 
def set_background():
    for ext in ['png', 'jpg', 'jpeg', 'webp']:
        path = f"Background.{ext}"
        if os.path.exists(path):
            with open(path, "rb") as f:
                data = base64.b64encode(f.read()).decode()
            st.markdown(f"""
                <style>
                .stApp {{
                    background-image: url("data:image/{ext};base64,{data}");
                    background-size: cover;
                    background-position: center;
                    background-attachment: fixed;
                }}
                .block-container {{
                    background-color: rgba(255, 255, 255, 0.82);
                    border-radius: 12px;
                    padding: 1.5rem;
                }}
                .stButton > button[type="primary"] {{
                    background-color: #003366 !important;
                    border-color: #003366 !important;
                    color: white !important;
                }}
                </style>
            """, unsafe_allow_html=True)
            return

# 
# دوال Excel
#
def _excel_styles():
    return dict(
        header_fill  = PatternFill("solid", fgColor="1F4E79"),
        alt_fill     = PatternFill("solid", fgColor="D6E4F0"),
        white_fill   = PatternFill("solid", fgColor="FFFFFF"),
        green_fill   = PatternFill("solid", fgColor="E2EFDA"),
        bold_white   = Font(name="Arial", bold=True, color="FFFFFF", size=11),
        normal_font  = Font(name="Arial", size=10),
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True),
        thin_border  = Border(
            left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin',  color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
        )
    )

def _sh(ws, row, cols, s):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = s['header_fill']; cell.font = s['bold_white']
        cell.alignment = s['center_align']; cell.border = s['thin_border']

def _sd(ws, row, cols, s, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = s['alt_fill'] if alt else s['white_fill']
        cell.font = s['normal_font']
        cell.alignment = s['center_align']; cell.border = s['thin_border']

def _write_summary_sheet(wb, results_table, s):
    ws1 = wb.active
    ws1.title = "النتائج التراكمية"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "الجدول التراكمي النهائي - Type 4"
    ws1["A1"].font      = Font(name="Arial", bold=True, color="1F4E79", size=14)
    ws1["A1"].alignment = s['center_align']
    ws1["A1"].fill      = PatternFill("solid", fgColor="D6E4F0")
    ws1.row_dimensions[1].height = 30
    for i, h in enumerate(["المطار", "النتيجة الخام", "النتيجة النهائية (÷0.85)"], 1):
        ws1.cell(row=2, column=i, value=h)
    _sh(ws1, 2, 3, s)
    for r_idx, row in enumerate(results_table, start=3):
        ws1.cell(row=r_idx, column=1, value=row["المطار"])
        ws1.cell(row=r_idx, column=2, value=row["النتيجة الخام"])
        ws1.cell(row=r_idx, column=3, value=row["النتيجة النهائية (÷0.85)"])
        _sd(ws1, r_idx, 3, s, alt=(r_idx % 2 == 0))
        ws1.cell(row=r_idx, column=2).number_format = "0.0000%"
        ws1.cell(row=r_idx, column=3).number_format = "0.0000%"
        ws1.cell(row=r_idx, column=3).fill = s['green_fill']
        ws1.cell(row=r_idx, column=3).font = Font(name="Arial", bold=True, color="375623", size=10)
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 28

def _write_detail_sheets(wb, detail_data, s):
    for sheet_name, df_detail in detail_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.sheet_view.rightToLeft = True
        for c_idx, col_name in enumerate(df_detail.columns, 1):
            ws.cell(row=1, column=c_idx, value=col_name)
        _sh(ws, 1, len(df_detail.columns), s)
        for r_idx, row in enumerate(df_detail.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            _sd(ws, r_idx, len(df_detail.columns), s, alt=(r_idx % 2 == 0))
        for c_idx, col_name in enumerate(df_detail.columns, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = max(18, len(str(col_name)) + 4)

def build_excel_report(results_table, detail_data=None):
    s = _excel_styles(); wb = Workbook()
    _write_summary_sheet(wb, results_table, s)
    if detail_data: _write_detail_sheets(wb, detail_data, s)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

def fix_arabic(text):
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        return get_display(arabic_reshaper.reshape(str(text)))
    except Exception:
        return str(text)

def make_chart_main(df_main, airport_code):
    categories = [fix_arabic(c) for c in df_main['Category'].tolist()]
    values     = [float(v) * 100 for v in df_main['النتيجة النهائية للتصنيف'].tolist()]
    fig, ax = plt.subplots(figsize=(13, max(5, len(categories) * 0.55 + 2)))
    fig.patch.set_facecolor('#F7FBFF'); ax.set_facecolor('#F7FBFF')
    colors = ['#1F4E79' if v >= 70 else '#2E75B6' if v >= 50 else '#9DC3E6' for v in values]
    bars = ax.barh(categories, values, color=colors, edgecolor='white', height=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.2, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va='center', ha='left', fontsize=9, color='#1F4E79', fontweight='bold')
    ax.set_xlim(0, max(values) * 1.3 if values else 100)
    ax.set_title(fix_arabic(f'نتائج التصنيفات الرئيسية - {airport_code}'),
                 fontsize=13, fontweight='bold', color='#1F4E79', pad=15)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.4, color='#BFBFBF')
    from matplotlib.patches import Patch
    ax.legend(handles=[
        Patch(facecolor='#1F4E79', label=fix_arabic('ممتاز (≥ 70%)')),
        Patch(facecolor='#2E75B6', label=fix_arabic('جيد (≥ 50%)')),
        Patch(facecolor='#9DC3E6', label=fix_arabic('يحتاج تحسين (< 50%)')),
    ], loc='lower right', fontsize=8, framealpha=0.7)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); buf.seek(0)
    return buf

def make_chart_sub(df_sub, airport_code, category_name):
    df_cat = df_sub[df_sub['Category'] == category_name].copy()
    if df_cat.empty: return None
    labels = [fix_arabic(r['Subcategory']) for _, r in df_cat.iterrows()]
    values = [float(v) * 100 for v in df_cat['الوزن × الإنجاز'].tolist()]
    fig, ax = plt.subplots(figsize=(12, max(3.5, len(labels) * 0.5 + 1.5)))
    fig.patch.set_facecolor('#F7FBFF'); ax.set_facecolor('#F7FBFF')
    palette = plt.cm.Blues(np.linspace(0.45, 0.88, max(len(labels), 1)))
    bars = ax.barh(labels, values, color=palette, edgecolor='white', height=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va='center', ha='left', fontsize=9, color='#1F4E79')
    ax.set_xlim(0, max(values) * 1.3 if values else 100)
    ax.set_title(fix_arabic(f'{category_name} | {airport_code}'),
                 fontsize=11, fontweight='bold', color='#1F4E79', pad=10)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.4, color='#BFBFBF')
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); buf.seek(0)
    return buf

def build_excel_with_charts(results_table, detail_data):
    s = _excel_styles(); wb = Workbook()
    _write_summary_sheet(wb, results_table, s)
    if detail_data: _write_detail_sheets(wb, detail_data, s)
    airports_in_data = {k.replace("MainCat_", "") for k in detail_data if k.startswith("MainCat_")}
    if airports_in_data:
        ws_c = wb.create_sheet(title="الرسومات البيانية")
        ws_c.sheet_view.rightToLeft = True
        ws_c["A1"] = "الرسومات البيانية"
        ws_c["A1"].font = Font(name="Arial", bold=True, color="1F4E79", size=14)
        ws_c["A1"].fill = PatternFill("solid", fgColor="D6E4F0")
        ws_c["A1"].alignment = s['center_align']
        cur_row = 3
        for code in sorted(airports_in_data):
            ws_c.cell(row=cur_row, column=1, value=f"مطار {code}")
            ws_c.cell(row=cur_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
            ws_c.cell(row=cur_row, column=1).fill = PatternFill("solid", fgColor="1F4E79")
            ws_c.cell(row=cur_row, column=1).alignment = s['center_align']
            ws_c.merge_cells(f"A{cur_row}:P{cur_row}"); cur_row += 1
            if f"MainCat_{code}" in detail_data:
                df_m = detail_data[f"MainCat_{code}"]
                buf_m = make_chart_main(df_m, code)
                img_m = XLImage(buf_m)
                img_m.width = 750; img_m.height = max(320, len(df_m) * 28 + 80)
                ws_c.add_image(img_m, f"A{cur_row}")
                cur_row += max(22, int(img_m.height / 15) + 2)
            if f"SubCat_{code}" in detail_data:
                df_s = detail_data[f"SubCat_{code}"]
                ws_c.cell(row=cur_row, column=1, value=f"التصنيفات الفرعية - {code}")
                ws_c.cell(row=cur_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                ws_c.cell(row=cur_row, column=1).fill = PatternFill("solid", fgColor="2E75B6")
                ws_c.cell(row=cur_row, column=1).alignment = s['center_align']
                ws_c.merge_cells(f"A{cur_row}:P{cur_row}"); cur_row += 1
                for cat in df_s['Category'].unique():
                    buf_s = make_chart_sub(df_s, code, cat)
                    if buf_s is None: continue
                    img_s = XLImage(buf_s)
                    img_s.width = 720; img_s.height = max(220, len(df_s[df_s['Category'] == cat]) * 32 + 90)
                    ws_c.add_image(img_s, f"A{cur_row}")
                    cur_row += max(16, int(img_s.height / 15) + 2)
            cur_row += 3
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out


# 
# دالة كشف علامة * في بيانات التقييم
# 
def _detect_star_values(df, columns_to_check):
    """تفحص الأعمدة المحددة وتعرض تنبيه إذا وُجدت علامة * بدلاً من رقم"""
    star_found = False
    star_msgs = []
    for col_name in columns_to_check:
        if col_name not in df.columns:
            continue
        star_mask = df[col_name].astype(str).str.contains(r'\*', na=False)
        star_rows = df[star_mask]
        if not star_rows.empty:
            star_found = True
            # محاولة عرض أرقام العناصر
            if 'الرقم القديم' in df.columns:
                refs = star_rows['الرقم القديم'].dropna()
                refs_list = [str(int(x)) if pd.notna(x) and x == int(x) else str(x) for x in refs]
                star_msgs.append(f"**{col_name}**: العناصر رقم ({', '.join(refs_list)}) — عدد: {len(star_rows)}")
            else:
                star_msgs.append(f"**{col_name}**: عدد {len(star_rows)} صف يحتوي على *")
    if star_found:
        st.warning(
            "⚠️ **تنبيه:** تم العثور على علامة ( * ) بدلاً من رقم في الحقول التالية:\n\n"
            + "\n\n".join(star_msgs)
            + "\n\n"
            + "هذه العناصر سيتم تجاهلها في الحساب (تُعامل كقيم فارغة)."
        )
    return star_found


# 
# دوال الحساب (المنطق ما يتغير)
# 
def prepare_raw_data(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    
    old_col = None
    possible_names = ['الرقم القديم', 'الرقم', 'رقم المعيار', 'ref.', 'ref', 'id', 'item']
    for p in possible_names:
        for col in df.columns:
            if p.lower() == col.lower(): old_col = col; break
        if old_col: break
        
    if not old_col:
        for col in df.columns:
            if 'رقم' in col and 'قديم' in col: old_col = col; break
            
    if not old_col and len(df.columns) > 1: old_col = df.columns[1]
        
    if old_col != 'الرقم القديم': df.rename(columns={old_col: 'الرقم القديم'}, inplace=True)
    df['الرقم القديم'] = pd.to_numeric(df['الرقم القديم'], errors='coerce')
    df['Link_ID'] = df['الرقم القديم'].apply(lambda x: f"{int(x)}&4" if pd.notna(x) else '')
    
    j_col = None
    for col in df.columns:
        if col == 'التقييم الداخلي/الأساسي': j_col = col; break
        if ('داخلي' in col or 'أساسي' in col or 'اساسي' in col) and 'ملاحظات' not in col: j_col = col; break
    if not j_col and len(df.columns) > 8: j_col = df.columns[8]
    if j_col and j_col != 'التقييم الداخلي/الأساسي': df.rename(columns={j_col: 'التقييم الداخلي/الأساسي'}, inplace=True)
        
    k_col = None
    for col in df.columns:
        if col == 'التقييم الدولي': k_col = col; break
        if 'دولي' in col and 'ملاحظات' not in col: k_col = col; break
    if not k_col and len(df.columns) > 9: k_col = df.columns[9]
    if k_col and k_col != 'التقييم الدولي': df.rename(columns={k_col: 'التقييم الدولي'}, inplace=True)
        
    return df

def _aggregate_type1_style(merged, display_sub, display_main):
    merged_audited = merged.dropna(subset=['Final_Soure']).copy()
    
    sub_agg = merged_audited.groupby(['التصنيف الصحيح', 'التصنيف الفرعي الصحيح']).agg(
        Item_Score_Sum=('Item_Score', 'sum'),
        Category_score_sum=('وزن العنصر', 'sum')
    ).reset_index()
    
    sub_agg['Subcategory_Score'] = sub_agg['Item_Score_Sum'] / sub_agg['Category_score_sum']

    sub_agg['_cat_norm'] = sub_agg['التصنيف الصحيح'].apply(normalize_arabic)
    sub_agg['_subcat_norm'] = sub_agg['التصنيف الفرعي الصحيح'].apply(normalize_arabic)

    sub_final = pd.merge(display_sub, sub_agg[['_cat_norm', '_subcat_norm', 'Subcategory_Score']], on=['_cat_norm', '_subcat_norm'], how='left')
    sub_final['Subcategory_Score'] = sub_final['Subcategory_Score'].fillna(0)
    sub_final['Category_score'] = sub_final['Weight'] * sub_final['Subcategory_Score']

    main_agg = sub_final.groupby('Category')['Category_score'].sum().reset_index()
    main_final = pd.merge(display_main, main_agg, on='Category', how='left').fillna(0)
    main_final['Final_Calc'] = main_final['Category_score'] * main_final['Weight']

    raw_score = main_final['Final_Calc'].sum()
    final_score = raw_score / 0.85

    return sub_final, main_final, raw_score, final_score

def calc_single_terminal(df_raw, lookup_df, display_sub, display_main, dom_pax, int_pax):
    df_raw = prepare_raw_data(df_raw)
    
    # ── كشف علامة * في بيانات التقييم ──
    _detect_star_values(df_raw, ['التقييم الداخلي/الأساسي', 'التقييم الدولي'])
    
    df_raw['J_score'] = pd.to_numeric(df_raw['التقييم الداخلي/الأساسي'], errors='coerce') / 2
    df_raw['K_score'] = pd.to_numeric(df_raw['التقييم الدولي'], errors='coerce') / 2

    merged = pd.merge(lookup_df, df_raw[['Link_ID', 'J_score', 'K_score']], left_on='Ref.', right_on='Link_ID', how='left')
    merged['وزن العنصر'] = pd.to_numeric(merged['وزن العنصر'], errors='coerce').fillna(0)

    def calc_idk_single(row):
        num, den = 0, 0
        if pd.notna(row['J_score']):
            num += row['J_score'] * dom_pax
            den += dom_pax
        if pd.notna(row['K_score']):
            num += row['K_score'] * int_pax
            den += int_pax
        return num / den if den > 0 else np.nan

    merged['Final_Soure'] = merged.apply(calc_idk_single, axis=1)
    merged['Item_Score'] = merged['Final_Soure'] * merged['وزن العنصر']

    sub_final, main_final, raw_score, final_score = _aggregate_type1_style(merged, display_sub, display_main)
    return merged, (sub_final, main_final, raw_score, final_score)

def _score_terminal(df, dom_pax, int_pax, t_type, score_col):
    j = pd.to_numeric(df['التقييم الداخلي/الأساسي'], errors='coerce') / 2
    k_raw = df['التقييم الدولي'] if 'التقييم الدولي' in df.columns else pd.Series(np.nan, index=df.index, dtype=float)
    k = pd.to_numeric(k_raw, errors='coerce') / 2

    if t_type == "داخلي + دولي":
        total = dom_pax + int_pax
        num = j.fillna(0) * dom_pax + k.fillna(0) * int_pax
        den = j.notna().astype(float) * dom_pax + k.notna().astype(float) * int_pax
        df[score_col] = num / den.replace(0, np.nan) if total > 0 else j
        return total
    elif t_type == "داخلي فقط":
        df[score_col] = j
        return dom_pax
    else:
        df[score_col] = k
        return int_pax


def calc_multi_terminal(df1, df2, lookup_df, display_sub, display_main,
                        t1_dom_pax, t1_int_pax, t1_type,
                        t2_dom_pax, t2_int_pax, t2_type):
    df1 = prepare_raw_data(df1)
    df2 = prepare_raw_data(df2)

    st.markdown("**📄 ملف Terminal 1:**")
    _detect_star_values(df1, ['التقييم الداخلي/الأساسي', 'التقييم الدولي'])
    st.markdown("**📄 ملف Terminal 2:**")
    _detect_star_values(df2, ['التقييم الداخلي/الأساسي', 'التقييم الدولي'])

    t1_total = _score_terminal(df1, t1_dom_pax, t1_int_pax, t1_type, 'score_T1')
    t2_total = _score_terminal(df2, t2_dom_pax, t2_int_pax, t2_type, 'score_T2')

    df1_scores = df1.drop_duplicates('Link_ID')[['Link_ID', 'score_T1']]
    df2_scores = df2.drop_duplicates('Link_ID')[['Link_ID', 'score_T2']]

    merged = pd.merge(lookup_df, df1_scores, left_on='Ref.', right_on='Link_ID', how='left')
    merged = pd.merge(merged, df2_scores, left_on='Ref.', right_on='Link_ID', how='left')
    merged['وزن العنصر'] = pd.to_numeric(merged['وزن العنصر'], errors='coerce').fillna(0)

    def calc_idk_skip(row):
        num, den = 0, 0
        if pd.notna(row['score_T1']) and t1_total > 0:
            num += row['score_T1'] * t1_total; den += t1_total
        if pd.notna(row['score_T2']) and t2_total > 0:
            num += row['score_T2'] * t2_total; den += t2_total
        return num / den if den > 0 else np.nan

    merged['Final_Soure'] = merged.apply(calc_idk_skip, axis=1)
    merged['Item_Score']  = merged['Final_Soure'] * merged['وزن العنصر']

    sub_final, main_final, raw_score, final_score = _aggregate_type1_style(merged, display_sub, display_main)
    return merged, (sub_final, main_final, raw_score, final_score)


# 
# الدالة الرئيسية
# 
def run_type_4():
    set_background()
    st.markdown("<div style='margin-top: 80px;'></div>", unsafe_allow_html=True)
    st.header(" Type 4  ")

    if 'final_results_table_4' not in st.session_state: st.session_state.final_results_table_4 = []
    if 'detail_data_4' not in st.session_state: st.session_state.detail_data_4 = {}

    # ── [1] ملف الأوزان ──
    st.subheader("📂 الخطوة 1: رفع ملف الأوزان")
    weights_file = st.file_uploader("ارفع ملف الأوزان", type=['xlsx'], key="weights_uploader_4")
    if not weights_file:
        st.info(" ارفع ملف الأوزان أولاً للمتابعة💡")
        return

    try:
        w_df = pd.read_excel(weights_file, sheet_name='Weights')
        w_df.columns = w_df.columns.str.strip()
        w4 = w_df[w_df['Type']==4][['Ref.','التصنيف','التصنيف الفرعي','weight of item']].copy()
        w4.columns = ['Ref.','التصنيف الصحيح','التصنيف الفرعي الصحيح','وزن العنصر']
        w4['Ref.'] = w4['Ref.'].astype(str).str.strip()
        
        sub_df = pd.read_excel(weights_file, sheet_name='subcategory weight')
        sub_df.columns = sub_df.columns.str.strip()
        yc = 'YEAR' if 'YEAR' in sub_df.columns else 'YAER'
        display_sub = sub_df[(sub_df['Type']==4)&(sub_df[yc]==2024)][['Category','Subcategory','Weight']].copy()
        display_sub['_cat_norm']    = display_sub['Category'].apply(normalize_arabic)
        display_sub['_subcat_norm'] = display_sub['Subcategory'].apply(normalize_arabic)

        main_df = pd.read_excel(weights_file, sheet_name='main category weight')
        main_df.columns = main_df.columns.str.strip()
        ycm = 'YEAR' if 'YEAR' in main_df.columns else 'YAER'
        display_main = main_df[(main_df['Type']==4)&(main_df[ycm]==2024)][['Category','Weight']].copy()
        display_main['_cat_norm'] = display_main['Category'].apply(normalize_arabic)
    except Exception as e:
        st.error(f"❌ خطأ في ملف الأوزان: {e}")
        return

    st.success(" تم تحميل ملف الأوزان ✔")
    c1, c2 = st.columns(2)
    with c1:
        with st.expander("أوزان التصنيفات الفرعية", expanded=False):
            st.dataframe(display_sub[['Category','Subcategory','Weight']], use_container_width=True)
    with c2:
        with st.expander("أوزان التصنيفات الرئيسية", expanded=False):
            st.dataframe(display_main[['Category','Weight']], use_container_width=True)
    st.divider()

    # ── [2] ملف الباسنجر ──
    st.subheader("👥 الخطوة 2: رفع ملف بيانات الركاب")
    pas_file = st.file_uploader("ارفع ملف Airport Data", type=['xlsx','csv'], key="pas_uploader_4")
    if not pas_file:
        st.info(" ارفع ملف الركاب للمتابعة💡")
        return

    try:
        df_pas = pd.read_csv(pas_file) if pas_file.name.endswith('.csv') else pd.read_excel(pas_file)
        df_pas.columns = df_pas.columns.str.strip()
        for col in ['Domestic Arrival','Domestic Departure','International Arrival Total','International Departure Total']:
            if col in df_pas.columns: df_pas[col] = pd.to_numeric(df_pas[col], errors='coerce').fillna(0)
        st.session_state['pas_data_4'] = df_pas
    except Exception as e:
        st.error(f"❌ خطأ في ملف الركاب: {e}")
        return

    st.divider()

    # ── [3] نوع المطار ──
    st.subheader("🏢 الخطوة 3: بيانات المطار")
    airport_name = st.text_input("✈️ اسم/رمز المطار", placeholder="مثال: AJF أو TIF")
    if not airport_name: return

    st.write("🔀 نوع المطار:")
    col_s, col_m, col_sm = st.columns(3)
    airport_type_key = f"airport_type_4_{airport_name}"
    if airport_type_key not in st.session_state: st.session_state[airport_type_key] = None

    with col_s:
        if st.button("🛬 صالة وحدة", use_container_width=True, type="primary" if st.session_state[airport_type_key] == "single" else "secondary", key=f"btn_single_{airport_name}"):
            st.session_state[airport_type_key] = "single"
            st.rerun()
    with col_m:
        if st.button("🛬🛬 صالتان (ملفين)", use_container_width=True, type="primary" if st.session_state[airport_type_key] == "multi" else "secondary", key=f"btn_multi_{airport_name}"):
            st.session_state[airport_type_key] = "multi"
            st.rerun()
    with col_sm:
        if st.button("🛬🛬 صالتان (ملف واحد)", use_container_width=True, type="primary" if st.session_state[airport_type_key] == "single_multi_pax" else "secondary", key=f"btn_single_multi_{airport_name}"):
            st.session_state[airport_type_key] = "single_multi_pax"
            st.rerun()

    airport_type = st.session_state[airport_type_key]
    if not airport_type: return

    # ── [4] رفع الملفات والحساب ──
    df_pas = st.session_state['pas_data_4']
    _available_terminals_4 = sorted(df_pas['Airport Terminal'].dropna().unique().tolist()) if 'Airport Terminal' in df_pas.columns else []

    if airport_type == "single":
        st.info("📋 صالة وحدة — فيه تقييم داخلي (J) ودولي (K)")
        terminal_input = st.selectbox("اختر الـ Terminal من ملف الباسنجر", options=[""] + _available_terminals_4, key=f"terminal_single_{airport_name}")
        raw_file = st.file_uploader("📄 ارفع ملف Raw Data", type=['xlsx','csv'], key=f"raw_single_{airport_name}")

        if terminal_input and raw_file:
            rows = df_pas[df_pas['Airport Terminal'] == terminal_input]
            if rows.empty:
                st.error(f"❌ ما وجدت الـ Terminal: {terminal_input}")
            else:
                dom_pax = int(rows['Domestic Arrival'].sum() + rows['Domestic Departure'].sum())
                int_pax = int(rows['International Arrival Total'].sum() + rows['International Departure Total'].sum())
                with st.expander("📅 مقارنة مع السنة السابقة (اختياري)"):
                    _ps1, _ps2 = st.columns(2)
                    prev_raw_s   = _ps1.number_input("النتيجة الخام للسنة السابقة",   min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key=f"prev_raw_s_{airport_name}")
                    prev_final_s = _ps2.number_input("النتيجة النهائية للسنة السابقة", min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key=f"prev_final_s_{airport_name}")
                if st.button(" ابدأ الحساب 🔎", type="primary", key=f"calc_single_{airport_name}"):
                    with st.spinner(" جاري الحساب ⏳"):
                        df_raw = pd.read_csv(raw_file) if raw_file.name.endswith('.csv') else pd.read_excel(raw_file)
                        merged, (sub_final, main_final, raw_score, final_score) = calc_single_terminal(
                            df_raw, w4, display_sub, display_main, dom_pax, int_pax
                        )
                        _save_and_show(airport_name, raw_score, final_score, merged, sub_final, main_final, airport_type, prev_raw_s, prev_final_s)

    elif airport_type == "single_multi_pax":
        st.info("📋 ملف تقييم واحد (J=داخلي، K=دولي) — لكن بيانات الركاب موزعة على صالتين في ملف الباسنجر")

        col_t1, col_t2 = st.columns(2)
        with col_t1:
            terminal_dom = st.selectbox("Terminal الداخلي في ملف الباسنجر", options=[""] + _available_terminals_4, key=f"t_dom_{airport_name}")
        with col_t2:
            terminal_int = st.selectbox("Terminal الدولي في ملف الباسنجر", options=[""] + _available_terminals_4, key=f"t_int_{airport_name}")

        raw_file = st.file_uploader("📄 ارفع ملف Raw Data", type=['xlsx','csv'], key=f"raw_sm_{airport_name}")

        if terminal_dom and terminal_int and raw_file:
            rows_dom = df_pas[df_pas['Airport Terminal'] == terminal_dom]
            rows_int = df_pas[df_pas['Airport Terminal'] == terminal_int]

            if rows_dom.empty and rows_int.empty:
                st.error(f"❌ ما وجدت أي من الـ Terminals")
            elif rows_dom.empty:
                st.error(f"❌ ما وجدت Terminal الداخلي: {terminal_dom}")
            elif rows_int.empty:
                st.error(f"❌ ما وجدت Terminal الدولي: {terminal_int}")
            else:
                dom_pax = int(rows_dom['Domestic Arrival'].sum() + rows_dom['Domestic Departure'].sum()
                            + rows_int['Domestic Arrival'].sum() + rows_int['Domestic Departure'].sum())
                int_pax = int(rows_dom['International Arrival Total'].sum() + rows_dom['International Departure Total'].sum()
                            + rows_int['International Arrival Total'].sum() + rows_int['International Departure Total'].sum())

                st.caption(f"👥 إجمالي الركاب — داخلي: **{dom_pax:,}** | دولي: **{int_pax:,}**")
                with st.expander("📅 مقارنة مع السنة السابقة (اختياري)"):
                    _psm1, _psm2 = st.columns(2)
                    prev_raw_sm   = _psm1.number_input("النتيجة الخام للسنة السابقة",   min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key=f"prev_raw_sm_{airport_name}")
                    prev_final_sm = _psm2.number_input("النتيجة النهائية للسنة السابقة", min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key=f"prev_final_sm_{airport_name}")
                if st.button(" ابدأ الحساب 🔎", type="primary", key=f"calc_sm_{airport_name}"):
                    with st.spinner(" جاري الحساب ⏳"):
                        df_raw = pd.read_csv(raw_file) if raw_file.name.endswith('.csv') else pd.read_excel(raw_file)
                        merged, (sub_final, main_final, raw_score, final_score) = calc_single_terminal(
                            df_raw, w4, display_sub, display_main, dom_pax, int_pax
                        )
                        _save_and_show(airport_name, raw_score, final_score, merged, sub_final, main_final, "single", prev_raw_sm, prev_final_sm)

    else:  # multi
        st.info("📋 صالتان — كل صالة لها ملف Raw Data منفصل")
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            terminal_1 = st.selectbox("Terminal 1 في ملف الباسنجر", options=[""] + _available_terminals_4, key=f"t1_name_{airport_name}")
            t1_type = st.radio("نوع بيانات Terminal 1", ["داخلي + دولي", "داخلي فقط", "دولي فقط"], key=f"t1_type_{airport_name}", horizontal=True)
            raw_t1 = st.file_uploader("📄 ملف Raw Data - Terminal 1", type=['xlsx','csv'], key=f"raw_t1_{airport_name}")
        with col_t2:
            terminal_2 = st.selectbox("Terminal 2 في ملف الباسنجر", options=[""] + _available_terminals_4, key=f"t2_name_{airport_name}")
            t2_type = st.radio("نوع بيانات Terminal 2", ["داخلي + دولي", "داخلي فقط", "دولي فقط"], key=f"t2_type_{airport_name}", horizontal=True)
            raw_t2 = st.file_uploader("📄 ملف Raw Data - Terminal 2", type=['xlsx','csv'], key=f"raw_t2_{airport_name}")

        if terminal_1 and terminal_2 and raw_t1 and raw_t2:
            rows_t1 = df_pas[df_pas['Airport Terminal'] == terminal_1]
            rows_t2 = df_pas[df_pas['Airport Terminal'] == terminal_2]

            if rows_t1.empty or rows_t2.empty:
                st.error("❌ تأكد من أسماء الـ Terminals")
            else:
                t1_dom = int(rows_t1['Domestic Arrival'].sum() + rows_t1['Domestic Departure'].sum())
                t1_int = int(rows_t1['International Arrival Total'].sum() + rows_t1['International Departure Total'].sum())
                t2_dom = int(rows_t2['Domestic Arrival'].sum() + rows_t2['Domestic Departure'].sum())
                t2_int = int(rows_t2['International Arrival Total'].sum() + rows_t2['International Departure Total'].sum())

                t1_used = t1_dom + t1_int if t1_type == "داخلي + دولي" else t1_dom if t1_type == "داخلي فقط" else t1_int
                t2_used = t2_dom + t2_int if t2_type == "داخلي + دولي" else t2_dom if t2_type == "داخلي فقط" else t2_int
                st.caption(f"👥 Terminal 1 ({t1_type}): **{t1_used:,}** | Terminal 2 ({t2_type}): **{t2_used:,}**")

                with st.expander("📅 مقارنة مع السنة السابقة (اختياري)"):
                    _pm1, _pm2 = st.columns(2)
                    prev_raw_m   = _pm1.number_input("النتيجة الخام للسنة السابقة",   min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key=f"prev_raw_m_{airport_name}")
                    prev_final_m = _pm2.number_input("النتيجة النهائية للسنة السابقة", min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key=f"prev_final_m_{airport_name}")
                if st.button(" ابدأ الحساب 🔎", type="primary", key=f"calc_multi_{airport_name}"):
                    with st.spinner(" جاري الحساب ⏳"):
                        df1 = pd.read_csv(raw_t1) if raw_t1.name.endswith('.csv') else pd.read_excel(raw_t1)
                        df2 = pd.read_csv(raw_t2) if raw_t2.name.endswith('.csv') else pd.read_excel(raw_t2)

                        merged, (sub_final, main_final, raw_score, final_score) = calc_multi_terminal(
                            df1, df2, w4, display_sub, display_main,
                            t1_dom, t1_int, t1_type,
                            t2_dom, t2_int, t2_type
                        )
                        _save_and_show(airport_name, raw_score, final_score, merged, sub_final, main_final, airport_type, prev_raw_m, prev_final_m)

    # ── الجدول التراكمي والتحميل ──
    if st.session_state.final_results_table_4:
        st.divider()
        st.subheader("🏆 الجدول النهائي لجميع المطارات")
        _cum_disp_4 = pd.DataFrame(st.session_state.final_results_table_4).copy()
        _cum_disp_4['النتيجة الخام'] = _cum_disp_4['النتيجة الخام'] * 100
        _cum_disp_4['النتيجة النهائية (÷0.85)'] = _cum_disp_4['النتيجة النهائية (÷0.85)'] * 100
        st.dataframe(_cum_disp_4, use_container_width=True,
            column_config={
                'النتيجة الخام':            st.column_config.NumberColumn(format="%.4f%%"),
                'النتيجة النهائية (÷0.85)': st.column_config.NumberColumn(format="%.4f%%"),
            })

        st.divider()
        st.subheader("⬇️ تحميل النتائج")
        dl1, dl2, dl3 = st.columns(3)
        with dl1:
            st.markdown("**📊 Excel الكامل**")
            st.download_button("📥 تحميل Excel", data=build_excel_report(st.session_state.final_results_table_4, st.session_state.detail_data_4), file_name="Type4_Full.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with dl2:
            st.markdown("** Excel بدون رسومات**")
            st.download_button("📥 تحميل Excel", data=build_excel_report(st.session_state.final_results_table_4), file_name="Type4_Summary.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with dl3:
            st.markdown("**📈 Excel مع رسومات**")
            if st.session_state.detail_data_4:
                excel_charts = build_excel_with_charts(st.session_state.final_results_table_4, st.session_state.detail_data_4)
                st.download_button("📥 تحميل Excel + رسومات", data=excel_charts, file_name="Type4_Charts.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        
        st.divider()
        if st.button("🗑️ مسح جميع النتائج"):
            st.session_state.final_results_table_4 = []; st.session_state.detail_data_4 = {}; st.rerun()


def _save_and_show(airport_name, raw_score, final_score, merged, sub_final, main_final, airport_type, prev_raw=0.0, prev_final=0.0):
    use_prev = prev_raw > 0 or prev_final > 0
    st.session_state.final_results_table_4.append({"المطار": airport_name, "النتيجة الخام": round(raw_score, 6), "النتيجة النهائية (÷0.85)": round(final_score, 6)})

    disp_sub = sub_final[['Category','Subcategory','Weight','Subcategory_Score','Category_score']].copy()
    disp_sub.columns = ['Category','Subcategory','الوزن','الإنجاز','الوزن × الإنجاز']

    disp_main = main_final[['Category','Weight','Category_score','Final_Calc']].copy()
    disp_main.columns = ['Category','الوزن','مجموع التصنيف','النتيجة النهائية للتصنيف']

    st.session_state.detail_data_4[f"SubCat_{airport_name}"]  = disp_sub
    st.session_state.detail_data_4[f"MainCat_{airport_name}"] = disp_main

    st.success(f" تم الحساب بنجاح ✔")
    c1, c2 = st.columns(2)
    _delta_raw   = f"{(raw_score - prev_raw)*100:+.4f}%"   if use_prev else None
    _delta_final = f"{(final_score - prev_final)*100:+.4f}%" if use_prev else None
    c1.metric(" النتيجة الخام📊", f"{raw_score*100:.4f}%", delta=_delta_raw)
    c2.metric(" النتيجة النهائية (÷ 0.85)🏆", f"{final_score*100:.4f}%", delta=_delta_final)
    st.divider()

    if airport_type == "single":
        cols = ['Ref.', 'التصنيف الصحيح', 'J_score', 'K_score', 'Final_Soure', 'وزن العنصر', 'Item_Score']
    else:
        cols = ['Ref.', 'التصنيف الصحيح', 'score_T1', 'score_T2', 'Final_Soure', 'وزن العنصر', 'Item_Score']

    avail_cols = [c for c in cols if c in merged.columns]
    with st.expander("📋 تفاصيل العناصر (Item Level)", expanded=False): st.dataframe(merged[avail_cols], use_container_width=True)

    _disp_sub_pct = disp_sub.copy()
    _disp_sub_pct['الإنجاز']      = _disp_sub_pct['الإنجاز'] * 100
    _disp_sub_pct['الوزن × الإنجاز'] = _disp_sub_pct['الوزن × الإنجاز'] * 100
    with st.expander("📋 تفاصيل التصنيف الفرعي (Subcategory)", expanded=True):
        st.dataframe(_disp_sub_pct, use_container_width=True,
            column_config={
                'الإنجاز':          st.column_config.NumberColumn(format="%.4f%%"),
                'الوزن × الإنجاز':  st.column_config.NumberColumn(format="%.4f%%"),
            })

    _disp_main_pct = disp_main.copy()
    _disp_main_pct['مجموع التصنيف']          = _disp_main_pct['مجموع التصنيف'] * 100
    _disp_main_pct['النتيجة النهائية للتصنيف'] = _disp_main_pct['النتيجة النهائية للتصنيف'] * 100
    with st.expander("📋 تفاصيل التصنيف الرئيسي (Main Category)", expanded=True):
        st.dataframe(_disp_main_pct, use_container_width=True,
            column_config={
                'مجموع التصنيف':            st.column_config.NumberColumn(format="%.4f%%"),
                'النتيجة النهائية للتصنيف': st.column_config.NumberColumn(format="%.4f%%"),
            })
