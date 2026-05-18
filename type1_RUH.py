import streamlit as st
import pandas as pd
import numpy as np
import re
import base64
import os
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# 
# دالة تطبيع النص العربي
# 
def normalize_arabic(text):
    if not isinstance(text, str):
        return text
    text = text.strip()
    text = re.sub(r'[أإآا]', 'ا', text)
    text = re.sub(r'[ىي]', 'ي', text)
    text = re.sub(r'[ةه]', 'ه', text)
    text = re.sub(r'[\u064B-\u065F]', '', text)
    return text

# 
# دالة تحميل الخلفية
# 
def set_background():
    for ext in ['png', 'jpg', 'jpeg', 'webp']:
        path = f"Background.{ext}"
        if os.path.exists(path):
            with open(path, "rb") as f:
                data = base64.b64encode(f.read()).decode()
            st.markdown(f"""
                <style>
                .stApp {{
                    background-image: url("data:image/{ext};base64,{data}");
                    background-size: cover;
                    background-position: center;
                    background-attachment: fixed;
                }}
                .block-container {{
                    background-color: rgba(255, 255, 255, 0.82);
                    border-radius: 12px;
                    padding: 1.5rem;
                }}
                .stButton > button[type="primary"] {{
                    background-color: #003366 !important;
                    border-color: #003366 !important;
                    color: white !important;
                }}
                </style>
            """, unsafe_allow_html=True)
            return

# 
# دالة حساب pax لكل Terminal
# 
def get_terminal_pax(df_pas, terminal_map=None):
    if terminal_map is None:
        terminal_map = {
            'T1': 'King Khalid International Airport>1',
            'T3': 'King Khalid International Airport>3',
            'T4': 'King Khalid International Airport>4',
            'T5': 'King Khalid International Airport>5',
        }
    for col in ['Domestic Arrival', 'Domestic Departure', 'International Arrival Total', 'International Departure Total']:
        if col in df_pas.columns:
            df_pas[col] = pd.to_numeric(df_pas[col], errors='coerce').fillna(0)

    pax = {}
    for t_name, t_code in terminal_map.items():
        rows = df_pas[df_pas['Airport Terminal'] == t_code]
        dom = rows['Domestic Arrival'].sum() + rows['Domestic Departure'].sum()
        intl = rows['International Arrival Total'].sum() + rows['International Departure Total'].sum()
        pax[t_name] = dom + intl
    return pax

# 
# دالة حساب score لصالة واحدة (Skip Logic)
# 
def calc_terminal_scores(df_raw):
    df = df_raw.copy()
    df.columns = df.columns.str.strip()
    df['J'] = pd.to_numeric(df['التقييم الداخلي/الأساسي'], errors='coerce')
    df['score'] = df['J'].apply(lambda j: j / 2 if pd.notna(j) else np.nan)
    return df[['الرقم القديم', 'score']].set_index('الرقم القديم')['score']


# =============================================
# دالة كشف علامة * في بيانات التقييم
# =============================================
def _detect_star_values(df, columns_to_check, terminal_label=None):
    """تفحص الأعمدة المحددة وتعرض تنبيه إذا وُجدت علامة * بدلاً من رقم"""
    star_found = False
    star_msgs = []
    for col_name in columns_to_check:
        if col_name not in df.columns:
            continue
        star_mask = df[col_name].astype(str).str.contains(r'\*', na=False)
        star_rows = df[star_mask]
        if not star_rows.empty:
            star_found = True
            if 'الرقم القديم' in df.columns:
                refs = star_rows['الرقم القديم'].dropna()
                refs_list = [str(int(x)) if pd.notna(x) and x == int(x) else str(x) for x in refs]
                star_msgs.append(f"**{col_name}**: العناصر رقم ({', '.join(refs_list)}) — عدد: {len(star_rows)}")
            else:
                star_msgs.append(f"**{col_name}**: عدد {len(star_rows)} صف يحتوي على *")
    if star_found:
        prefix = f"**{terminal_label}** — " if terminal_label else ""
        st.warning(
            f"⚠️ **تنبيه:** {prefix}تم العثور على علامة ( * ) بدلاً من رقم في الحقول التالية:\n\n"
            + "\n\n".join(star_msgs)
            + "\n\n"
            + "هذه العناصر سيتم تجاهلها في الحساب (تُعامل كقيم فارغة)."
        )
    return star_found


# =============================================
# دوال Excel
# =============================================
def _excel_styles():
    return dict(
        header_fill  = PatternFill("solid", fgColor="1F4E79"),
        alt_fill     = PatternFill("solid", fgColor="D6E4F0"),
        white_fill   = PatternFill("solid", fgColor="FFFFFF"),
        green_fill   = PatternFill("solid", fgColor="E2EFDA"),
        bold_white   = Font(name="Arial", bold=True, color="FFFFFF", size=11),
        normal_font  = Font(name="Arial", size=10),
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True),
        thin_border  = Border(
            left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin',  color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
        )
    )

def _sh(ws, row, cols, s):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = s['header_fill']; cell.font = s['bold_white']
        cell.alignment = s['center_align']; cell.border = s['thin_border']

def _sd(ws, row, cols, s, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = s['alt_fill'] if alt else s['white_fill']
        cell.font = s['normal_font']
        cell.alignment = s['center_align']; cell.border = s['thin_border']

def _write_summary_sheet(wb, results_table, s):
    ws1 = wb.active
    ws1.title = "النتائج التراكمية"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "الجدول التراكمي النهائي - Type 1 RUH"
    ws1["A1"].font      = Font(name="Arial", bold=True, color="1F4E79", size=14)
    ws1["A1"].alignment = s['center_align']
    ws1["A1"].fill      = PatternFill("solid", fgColor="D6E4F0")
    ws1.row_dimensions[1].height = 30
    for i, h in enumerate(["المطار", "النتيجة الخام", "النتيجة النهائية (÷0.85)"], 1):
        ws1.cell(row=2, column=i, value=h)
    _sh(ws1, 2, 3, s)
    for r_idx, row in enumerate(results_table, start=3):
        ws1.cell(row=r_idx, column=1, value=row["المطار"])
        ws1.cell(row=r_idx, column=2, value=row["النتيجة الخام"])
        ws1.cell(row=r_idx, column=3, value=row["النتيجة النهائية (÷0.85)"])
        _sd(ws1, r_idx, 3, s, alt=(r_idx % 2 == 0))
        ws1.cell(row=r_idx, column=2).number_format = "0.0000%"
        ws1.cell(row=r_idx, column=3).number_format = "0.0000%"
        ws1.cell(row=r_idx, column=3).fill = s['green_fill']
        ws1.cell(row=r_idx, column=3).font = Font(name="Arial", bold=True, color="375623", size=10)
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 28

def _write_detail_sheets(wb, detail_data, s):
    for sheet_name, df_detail in detail_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.sheet_view.rightToLeft = True
        for c_idx, col_name in enumerate(df_detail.columns, 1):
            ws.cell(row=1, column=c_idx, value=col_name)
        _sh(ws, 1, len(df_detail.columns), s)
        for r_idx, row in enumerate(df_detail.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            _sd(ws, r_idx, len(df_detail.columns), s, alt=(r_idx % 2 == 0))
        for c_idx, col_name in enumerate(df_detail.columns, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = max(18, len(str(col_name)) + 4)

def build_excel_report(results_table, detail_data=None):
    s = _excel_styles(); wb = Workbook()
    _write_summary_sheet(wb, results_table, s)
    if detail_data: _write_detail_sheets(wb, detail_data, s)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

def fix_arabic(text):
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        return get_display(arabic_reshaper.reshape(str(text)))
    except Exception:
        return str(text)

def make_chart_main(df_main, airport_code):
    categories = [fix_arabic(c) for c in df_main['Category'].tolist()]
    values     = [float(v) * 100 for v in df_main['Final_Calc'].tolist()]
    fig, ax = plt.subplots(figsize=(13, max(5, len(categories) * 0.55 + 2)))
    fig.patch.set_facecolor('#F7FBFF'); ax.set_facecolor('#F7FBFF')
    colors = ['#1F4E79' if v >= 70 else '#2E75B6' if v >= 50 else '#9DC3E6' for v in values]
    bars = ax.barh(categories, values, color=colors, edgecolor='white', height=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.2, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va='center', ha='left', fontsize=9, color='#1F4E79', fontweight='bold')
    ax.set_xlim(0, max(values) * 1.3 if values else 100)
    ax.set_title(fix_arabic(f'نتائج التصنيفات الرئيسية - {airport_code}'),
                 fontsize=13, fontweight='bold', color='#1F4E79', pad=15)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.4, color='#BFBFBF')
    from matplotlib.patches import Patch
    ax.legend(handles=[
        Patch(facecolor='#1F4E79', label=fix_arabic('ممتاز (≥ 70%)')),
        Patch(facecolor='#2E75B6', label=fix_arabic('جيد (≥ 50%)')),
        Patch(facecolor='#9DC3E6', label=fix_arabic('يحتاج تحسين (< 50%)')),
    ], loc='lower right', fontsize=8, framealpha=0.7)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); buf.seek(0)
    return buf

def make_chart_sub(df_sub, airport_code, category_name):
    df_cat = df_sub[df_sub['Category'] == category_name].copy()
    if df_cat.empty: return None
    labels = [fix_arabic(r['Subcategory']) for _, r in df_cat.iterrows()]
    values = [float(v) * 100 for v in df_cat['Category_score'].tolist()]
    fig, ax = plt.subplots(figsize=(12, max(3.5, len(labels) * 0.5 + 1.5)))
    fig.patch.set_facecolor('#F7FBFF'); ax.set_facecolor('#F7FBFF')
    palette = plt.cm.Blues(np.linspace(0.45, 0.88, max(len(labels), 1)))
    bars = ax.barh(labels, values, color=palette, edgecolor='white', height=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va='center', ha='left', fontsize=9, color='#1F4E79')
    ax.set_xlim(0, max(values) * 1.3 if values else 100)
    ax.set_title(fix_arabic(f'{category_name} | {airport_code}'),
                 fontsize=11, fontweight='bold', color='#1F4E79', pad=10)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.4, color='#BFBFBF')
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); buf.seek(0)
    return buf

def build_excel_with_charts(results_table, detail_data):
    s = _excel_styles(); wb = Workbook()
    _write_summary_sheet(wb, results_table, s)
    if detail_data: _write_detail_sheets(wb, detail_data, s)
    airports_in_data = {k.replace("MainCat_", "") for k in detail_data if k.startswith("MainCat_")}
    if airports_in_data:
        ws_c = wb.create_sheet(title="الرسومات البيانية")
        ws_c.sheet_view.rightToLeft = True
        ws_c["A1"] = "الرسومات البيانية"
        ws_c["A1"].font = Font(name="Arial", bold=True, color="1F4E79", size=14)
        ws_c["A1"].fill = PatternFill("solid", fgColor="D6E4F0")
        ws_c["A1"].alignment = s['center_align']
        cur_row = 3
        for code in sorted(airports_in_data):
            ws_c.cell(row=cur_row, column=1, value=f"مطار {code}")
            ws_c.cell(row=cur_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
            ws_c.cell(row=cur_row, column=1).fill = PatternFill("solid", fgColor="1F4E79")
            ws_c.cell(row=cur_row, column=1).alignment = s['center_align']
            ws_c.merge_cells(f"A{cur_row}:P{cur_row}"); cur_row += 1
            if f"MainCat_{code}" in detail_data:
                df_m = detail_data[f"MainCat_{code}"]
                buf_m = make_chart_main(df_m, code)
                img_m = XLImage(buf_m)
                img_m.width = 750; img_m.height = max(320, len(df_m) * 28 + 80)
                ws_c.add_image(img_m, f"A{cur_row}")
                cur_row += max(22, int(img_m.height / 15) + 2)
            if f"SubCat_{code}" in detail_data:
                df_s = detail_data[f"SubCat_{code}"]
                ws_c.cell(row=cur_row, column=1, value=f"التصنيفات الفرعية - {code}")
                ws_c.cell(row=cur_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                ws_c.cell(row=cur_row, column=1).fill = PatternFill("solid", fgColor="2E75B6")
                ws_c.cell(row=cur_row, column=1).alignment = s['center_align']
                ws_c.merge_cells(f"A{cur_row}:P{cur_row}"); cur_row += 1
                for cat in df_s['Category'].unique():
                    buf_s = make_chart_sub(df_s, code, cat)
                    if buf_s is None: continue
                    img_s = XLImage(buf_s)
                    img_s.width = 720; img_s.height = max(220, len(df_s[df_s['Category'] == cat]) * 32 + 90)
                    ws_c.add_image(img_s, f"A{cur_row}")
                    cur_row += max(16, int(img_s.height / 15) + 2)
            cur_row += 3
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out


# 
# الدالة الرئيسية
# 
def run_type_1_RUH():
    set_background()
    st.markdown("<div style='margin-top: 80px;'></div>", unsafe_allow_html=True)
    st.header(" Type 1 - الرياض (RUH) 📊 ")

    if 'final_results_table_1_RUH' not in st.session_state:
        st.session_state.final_results_table_1_RUH = []
    if 'detail_data_1_RUH' not in st.session_state:
        st.session_state.detail_data_1_RUH = {}
    if 'extra_terminals_ruh' not in st.session_state:
        st.session_state.extra_terminals_ruh = []

    # [1] ملف الأوزان
    st.subheader("📂 الخطوة 1: رفع ملف الأوزان")
    weights_file = st.file_uploader("ارفع ملف الأوزان", type=['xlsx'], key="weights_ruh")
    if not weights_file:
        st.info(" ارفع ملف الأوزان أولاً للمتابعة💡")
        return

    try:
        xl = pd.ExcelFile(weights_file)
        w_items = pd.read_excel(weights_file, sheet_name='Weights')
        w_items.columns = w_items.columns.str.strip()
        w1 = w_items[w_items['Type'] == 1].copy()
        lookup_df = w1[['Ref.', 'التصنيف', 'التصنيف الفرعي', 'weight of item']].copy()
        lookup_df.columns = ['Ref.', 'التصنيف الصحيح', 'التصنيف الفرعي الصحيح', 'وزن العنصر']
        lookup_df['Ref.'] = lookup_df['Ref.'].astype(str).str.strip()

        sub_df = pd.read_excel(weights_file, sheet_name='subcategory weight')
        sub_df.columns = sub_df.columns.str.strip()
        display_sub = sub_df[(sub_df['Type'] == 1) & (sub_df['YEAR' if 'YEAR' in sub_df.columns else 'YAER'] == 2024)].copy()
        display_sub['_cat_norm'] = display_sub['Category'].apply(normalize_arabic)
        display_sub['_subcat_norm'] = display_sub['Subcategory'].apply(normalize_arabic)

        main_df = pd.read_excel(weights_file, sheet_name='main category weight')
        main_df.columns = main_df.columns.str.strip()
        display_main = main_df[(main_df['Type'] == 1) & (main_df['YEAR' if 'YEAR' in main_df.columns else 'YAER'] == 2024)].copy()
        display_main['_cat_norm'] = display_main['Category'].apply(normalize_arabic)
    except Exception as e:
        st.error(f"❌ خطأ في ملف الأوزان: {e}"); return

    # [2] ملف الباسنجر
    st.subheader("👥 الخطوة 2: رفع ملف بيانات الباسنجر")
    pas_file = st.file_uploader("ارفع ملف Airport Data", type=['xlsx', 'csv'], key="pas_ruh")
    if not pas_file: return

    # بناء terminal_map بما فيها الصالات الإضافية
    terminal_map = {
        'T1': 'King Khalid International Airport>1',
        'T3': 'King Khalid International Airport>3',
        'T4': 'King Khalid International Airport>4',
        'T5': 'King Khalid International Airport>5',
    }
    for et in st.session_state.extra_terminals_ruh:
        terminal_map[et['name']] = et['code']

    df_pas_ruh = pd.read_csv(pas_file) if pas_file.name.endswith('.csv') else pd.read_excel(pas_file)
    df_pas_ruh.columns = df_pas_ruh.columns.str.strip()
    pax = get_terminal_pax(df_pas_ruh.copy(), terminal_map)
    total_pax_all = sum(pax.values())

    pax_cols = st.columns(len(pax))
    for col, (t_name, t_val) in zip(pax_cols, pax.items()):
        col.metric(f"🏢 {t_name}", f"{t_val:,}")

    # [3] ملفات الصالات
    st.subheader("📂 الخطوة 3: رفع ملفات الصالات")
    col_t1, col_t3 = st.columns(2)
    col_t4, col_t5 = st.columns(2)
    f1 = col_t1.file_uploader("🏢 صالة T1", type=['xlsx'], key="t1")
    f3 = col_t3.file_uploader("🏢 صالة T3", type=['xlsx'], key="t3")
    f4 = col_t4.file_uploader("🏢 صالة T4", type=['xlsx'], key="t4")
    f5 = col_t5.file_uploader("🏢 صالة T5", type=['xlsx'], key="t5")

    # ── إضافة صالة جديدة ──
    with st.expander("➕ إضافة صالة جديدة (اختياري)"):
        ec1, ec2, ec3 = st.columns([2, 3, 1])
        new_t_name = ec1.text_input("اسم الصالة (مثال: T6)", key="new_t_name_ruh").strip().upper()
        _available_terminals_ruh = sorted(df_pas_ruh['Airport Terminal'].dropna().unique().tolist()) if 'Airport Terminal' in df_pas_ruh.columns else []
        new_t_code = ec2.selectbox("كود الصالة في ملف الباسنجر", options=[""] + _available_terminals_ruh, key="new_t_code_ruh")
        if ec3.button("➕ إضافة", key="add_terminal_ruh"):
            existing = [et['name'] for et in st.session_state.extra_terminals_ruh]
            if not new_t_name:
                st.warning("أدخل اسم الصالة.")
            elif not new_t_code:
                st.warning("أدخل كود الصالة.")
            elif new_t_name in ['T1', 'T3', 'T4', 'T5']:
                st.warning("هذه الصالة موجودة مسبقاً.")
            elif new_t_name in existing:
                st.warning(f"الصالة {new_t_name} مضافة مسبقاً.")
            else:
                st.session_state.extra_terminals_ruh.append({'name': new_t_name, 'code': new_t_code})
                st.rerun()

        if st.session_state.extra_terminals_ruh:
            st.markdown("**الصالات المضافة:**")
            for et in list(st.session_state.extra_terminals_ruh):
                rc1, rc2 = st.columns([5, 1])
                rc1.markdown(f"- **{et['name']}** ← `{et['code']}`")
                if rc2.button("🗑️ حذف", key=f"del_{et['name']}_ruh"):
                    st.session_state.extra_terminals_ruh = [
                        x for x in st.session_state.extra_terminals_ruh if x['name'] != et['name']
                    ]
                    st.rerun()

    # رفع ملفات الصالات الإضافية
    extra_files = {}
    if st.session_state.extra_terminals_ruh:
        st.markdown("**📂 ملفات الصالات الإضافية:**")
        extra_cols = st.columns(min(len(st.session_state.extra_terminals_ruh), 3))
        for i, et in enumerate(st.session_state.extra_terminals_ruh):
            ef = extra_cols[i % 3].file_uploader(f"🏢 صالة {et['name']}", type=['xlsx'], key=f"extra_{et['name']}_ruh")
            extra_files[et['name']] = ef

    # قائمة كل الصالات والملفات
    all_terminals = ['T1', 'T3', 'T4', 'T5'] + [et['name'] for et in st.session_state.extra_terminals_ruh]
    all_files     = [f1, f3, f4, f5] + [extra_files.get(et['name']) for et in st.session_state.extra_terminals_ruh]

    # ── مقارنة مع السنة السابقة ──
    with st.expander("📅 مقارنة مع السنة السابقة (اختياري)"):
        py1, py2 = st.columns(2)
        prev_raw_ruh   = py1.number_input("النتيجة الخام للسنة السابقة",   min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key="prev_raw_ruh")
        prev_final_ruh = py2.number_input("النتيجة النهائية للسنة السابقة", min_value=0.0, max_value=1.0, value=0.0, step=0.0001, format="%.4f", key="prev_final_ruh")
    use_prev_ruh = prev_raw_ruh > 0 or prev_final_ruh > 0

    if all(all_files):
        if st.button(" ابدأ الحساب 🔎", type="primary"):
            with st.spinner(" جاري الحساب ⏳"):
                try:
                    # ── كشف علامة * في كل ملف صالة ──
                    for t_label, t_file in zip(all_terminals, all_files):
                        t_file.seek(0)
                        df_check = pd.read_excel(t_file)
                        df_check.columns = df_check.columns.str.strip()
                        _detect_star_values(df_check, ['التقييم الداخلي/الأساسي'], terminal_label=f"صالة {t_label}")
                        t_file.seek(0)

                    s_dict = {t: calc_terminal_scores(pd.read_excel(f)) for t, f in zip(all_terminals, all_files)}

                    df_base = pd.read_excel(f1).copy()
                    df_base['Link_ID'] = df_base['الرقم القديم'].astype(str).str.strip() + '&1'
                    merged = pd.merge(df_base[['الرقم القديم', 'Link_ID']], lookup_df, left_on='Link_ID', right_on='Ref.', how='inner')

                    for t in all_terminals:
                        merged[f'score_{t}'] = merged['الرقم القديم'].map(s_dict[t])

                    def calc_idk_skip(row):
                        num, den = 0, 0
                        for t in all_terminals:
                            if pd.notna(row[f'score_{t}']):
                                num += row[f'score_{t}'] * pax[t]
                                den += pax[t]
                        return num / den if den > 0 else np.nan

                    merged['Final_Soure'] = merged.apply(calc_idk_skip, axis=1)
                    merged['Item_Score'] = merged['Final_Soure'] * merged['وزن العنصر']

                    merged_audited = merged.dropna(subset=['Final_Soure']).copy()
                    sub_agg = merged_audited.groupby(['التصنيف الصحيح', 'التصنيف الفرعي الصحيح']).agg(
                        Item_Score_Sum=('Item_Score', 'sum'),
                        Category_score=('وزن العنصر', 'sum')
                    ).reset_index()
                    sub_agg['Subcategory_Score'] = sub_agg['Item_Score_Sum'] / sub_agg['Category_score']

                    sub_agg['_cat_norm'] = sub_agg['التصنيف الصحيح'].apply(normalize_arabic)
                    sub_agg['_subcat_norm'] = sub_agg['التصنيف الفرعي الصحيح'].apply(normalize_arabic)

                    sub_final = pd.merge(display_sub, sub_agg[['_cat_norm', '_subcat_norm', 'Subcategory_Score']], on=['_cat_norm', '_subcat_norm'], how='left')
                    sub_final['Subcategory_Score'] = sub_final['Subcategory_Score'].fillna(0)
                    sub_final['Category_score'] = sub_final['Weight'] * sub_final['Subcategory_Score']

                    main_agg = sub_final.groupby('Category')['Category_score'].sum().reset_index()
                    main_final = pd.merge(display_main, main_agg, on='Category', how='left').fillna(0)
                    main_final['Final_Calc'] = main_final['Category_score'] * main_final['Weight']

                    raw_score = main_final['Final_Calc'].sum()
                    final_score = raw_score / 0.85

                    st.success(" تم الحساب بنجاح ✔")
                    r1, r2 = st.columns(2)
                    _delta_raw_ruh   = f"{(raw_score - prev_raw_ruh)*100:.4f}%"   if use_prev_ruh else None
                    _delta_final_ruh = f"{(final_score - prev_final_ruh)*100:.4f}%" if use_prev_ruh else None
                    r1.metric(" النتيجة الخام📊", f"{raw_score*100:.4f}%", delta=_delta_raw_ruh)
                    r2.metric(" النتيجة النهائية (÷ 0.85)🏆", f"{final_score*100:.4f}%", delta=_delta_final_ruh)

                    score_cols = [f'score_{t}' for t in all_terminals]
                    with st.expander("📋 تفاصيل العناصر (Item Level)", expanded=False):
                        st.dataframe(merged[['الرقم القديم', 'التصنيف الصحيح'] + score_cols + ['Final_Soure', 'وزن العنصر', 'Item_Score']], use_container_width=True)

                    with st.expander("📋 تفاصيل التصنيف الفرعي (Subcategory)", expanded=True):
                        _sub_disp = sub_final[['Category', 'Subcategory', 'Weight', 'Subcategory_Score', 'Category_score']].copy()
                        _sub_disp['Subcategory_Score'] = _sub_disp['Subcategory_Score'] * 100
                        _sub_disp['Category_score']    = _sub_disp['Category_score'] * 100
                        st.dataframe(_sub_disp, use_container_width=True,
                            column_config={
                                'Subcategory_Score': st.column_config.NumberColumn(format="%.4f%%"),
                                'Category_score':    st.column_config.NumberColumn(format="%.4f%%"),
                            })

                    with st.expander("📋 تفاصيل التصنيف الرئيسي (Main Category)", expanded=True):
                        _main_disp = main_final[['Category', 'Weight', 'Category_score', 'Final_Calc']].copy()
                        _main_disp['Category_score'] = _main_disp['Category_score'] * 100
                        _main_disp['Final_Calc']     = _main_disp['Final_Calc'] * 100
                        st.dataframe(_main_disp, use_container_width=True,
                            column_config={
                                'Category_score': st.column_config.NumberColumn(format="%.4f%%"),
                                'Final_Calc':     st.column_config.NumberColumn(format="%.4f%%"),
                            })

                    # حفظ النتائج
                    st.session_state.final_results_table_1_RUH.append({
                        "المطار": "RUH",
                        "النتيجة الخام": round(raw_score, 6),
                        "النتيجة النهائية (÷0.85)": round(final_score, 6)
                    })

                    st.session_state.detail_data_1_RUH["Items_RUH"] = merged[['الرقم القديم', 'التصنيف الصحيح'] + score_cols + ['Final_Soure', 'وزن العنصر', 'Item_Score']].copy()
                    st.session_state.detail_data_1_RUH["SubCat_RUH"] = sub_final[['Category', 'Subcategory', 'Weight', 'Subcategory_Score', 'Category_score']].copy()
                    st.session_state.detail_data_1_RUH["MainCat_RUH"] = main_final[['Category', 'Weight', 'Category_score', 'Final_Calc']].copy()

                except Exception as e:
                    st.error(f"❌ خطأ في الحساب: {e}")
                    import traceback; st.code(traceback.format_exc())

    # ── خيارات التحميل ──
    if st.session_state.final_results_table_1_RUH:
        st.divider()
        st.subheader("🏆 الجدول النهائي التراكمي")
        _cum_disp = pd.DataFrame(st.session_state.final_results_table_1_RUH).copy()
        _cum_disp['النتيجة الخام'] = _cum_disp['النتيجة الخام'] * 100
        _cum_disp['النتيجة النهائية (÷0.85)'] = _cum_disp['النتيجة النهائية (÷0.85)'] * 100
        st.dataframe(_cum_disp, use_container_width=True,
            column_config={
                'النتيجة الخام':            st.column_config.NumberColumn(format="%.4f%%"),
                'النتيجة النهائية (÷0.85)': st.column_config.NumberColumn(format="%.4f%%"),
            })
        st.subheader("⬇️ تحميل النتائج")
        dl1, dl2 = st.columns(2)

        with dl1:
            st.markdown("**📊 Excel بدون رسومات**")
            st.download_button("📥 تحميل Excel",
                data=build_excel_report(st.session_state.final_results_table_1_RUH, st.session_state.detail_data_1_RUH),
                file_name="Type1_RUH_Full_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_full_ruh")

        with dl2:
            st.markdown("**📈 Excel مع رسومات**")
            if st.session_state.detail_data_1_RUH:
                with st.spinner("جاري إنشاء الرسومات..."):
                    excel_charts = build_excel_with_charts(
                        st.session_state.final_results_table_1_RUH,
                        st.session_state.detail_data_1_RUH)
                st.download_button("📥 تحميل Excel + رسومات",
                    data=excel_charts, file_name="Type1_RUH_Charts_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dl_charts_ruh")

        st.divider()
        if st.button("🗑️ مسح جميع النتائج", key="clear_ruh"):
            st.session_state.final_results_table_1_RUH = []
            st.session_state.detail_data_1_RUH = {}
            st.rerun()


if __name__ == "__main__":
    run_type_1_RUH()